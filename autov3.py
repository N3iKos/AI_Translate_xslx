#!/usr/bin/env python3
"""
Universal Excel Translator  v2.0
Powered by Gemini · Groq · Cerebras

Requires : Python 3.8+, openpyxl, requests
Install  : pip install openpyxl requests
"""

# ── Imports ───────────────────────────────────────────────────────────────────
import os, re, sys, json, time, glob, hmac, shutil, hashlib, signal, threading
try:
    from langdetect import detect as _langdetect
    _LANGDETECT_OK = True
except ImportError:
    _LANGDETECT_OK = False
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

try:
    import requests
except ImportError:
    sys.exit("ERROR: 'requests' not installed.  Run: pip install requests")
try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string, get_column_letter
except ImportError:
    sys.exit("ERROR: 'openpyxl' not installed.  Run: pip install openpyxl")


# ── ANSI Colors ───────────────────────────────────────────────────────────────
R   = "\033[0m"
B   = "\033[1m"
CY  = "\033[96m"
GR  = "\033[92m"
YL  = "\033[93m"
RD  = "\033[91m"
BL  = "\033[94m"
MG  = "\033[95m"
WH  = "\033[97m"
GY  = "\033[90m"
C2  = "\033[36m"

NAV_CMDS    = {"back", "b", "0"}
CANCEL_CMDS = {"cancel", "quit", "exit", "q"}


# ══════════════════════════════════════════════════════════════════════════════
#  VAULT SYSTEM  (pure stdlib — no external crypto deps)
# ══════════════════════════════════════════════════════════════════════════════
CONFIG_DIR  = os.path.join(os.path.expanduser("~"), ".config", "excel-translator")
CONFIG_FILE = os.path.join(CONFIG_DIR, "keys.enc")
_MAGIC      = b"EXLTR10\x00"   # 8-byte file magic


def _pbkdf2(password, salt, iterations=310_000, dklen=32):
    return hashlib.pbkdf2_hmac("sha256", password.encode(), salt, iterations, dklen)


def _aes_ctr_crypt(key, nonce, data):
    """AES-CTR emulation via SHA-256 keystream blocks (XOR cipher)."""
    out = bytearray()
    counter = 0
    for i in range(0, len(data), 32):
        ks = hashlib.sha256(key + nonce + counter.to_bytes(8, "big")).digest()
        chunk = data[i:i + 32]
        out.extend(b ^ k for b, k in zip(chunk, ks))
        counter += 1
    return bytes(out)


def _save_vault(data, password):
    """Encrypt and write the full vault dict to disk."""
    plaintext = json.dumps(data).encode()
    salt  = os.urandom(16)
    nonce = os.urandom(16)
    dk    = _pbkdf2(password, salt)
    ct    = _aes_ctr_crypt(dk, nonce, plaintext)
    tag   = hmac.new(dk, ct, hashlib.sha256).digest()
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(CONFIG_FILE, "wb") as f:
        f.write(_MAGIC + salt + nonce + tag + ct)


def _load_all_raw(password):
    """Decrypt vault. Returns dict on success, None on wrong password / corrupt / missing."""
    try:
        with open(CONFIG_FILE, "rb") as f:
            raw = f.read()
        if raw[:8] != _MAGIC:
            return None
        salt, nonce, tag, ct = raw[8:24], raw[24:40], raw[40:72], raw[72:]
        dk = _pbkdf2(password, salt)
        if not hmac.compare_digest(hmac.new(dk, ct, hashlib.sha256).digest(), tag):
            return None
        return json.loads(_aes_ctr_crypt(dk, nonce, ct).decode())
    except Exception:
        return None


def _save_keys(provider_key, keys, password):
    """Merge keys into vault. Returns False if vault exists but password is wrong."""
    if keys_file_exists():
        existing = _load_all_raw(password)
        if existing is None:
            return False
    else:
        existing = {}
    existing[provider_key] = keys
    _save_vault(existing, password)
    return True


def _load_keys(provider_key, password):
    """Returns list of keys, [] if provider not found, None if wrong password."""
    data = _load_all_raw(password)
    if data is None:
        return None
    return data.get(provider_key, [])


def _delete_keys(provider_key, password):
    data = _load_all_raw(password)
    if data is None or provider_key not in data:
        return False
    del data[provider_key]
    _save_vault(data, password)
    return True


def keys_file_exists():
    return os.path.isfile(CONFIG_FILE)


# ══════════════════════════════════════════════════════════════════════════════
#  TRANSLATION PRESETS
# ══════════════════════════════════════════════════════════════════════════════
PRESETS = PRESETS = [
    {
        "key": "formal",
        "label": "Formal / Professional",
        "prompt": (
            "Translate the text into a formal and professional tone.\n"
            "Requirements:\n"
            "- Use correct grammar and professional vocabulary.\n"
            "- Avoid slang, contractions, or casual expressions.\n"
            "- Preserve the original meaning exactly.\n"
            "- Do not add, remove, or reinterpret information.\n"
            "- Keep numbers, names, and formatting unchanged.\n"
            "- Output ONLY the translated text."
        ),
    },
    {
        "key": "casual",
        "label": "Natural / Casual",
        "prompt": (
            "Translate the text in a natural and conversational tone.\n"
            "Requirements:\n"
            "- Use everyday language that sounds natural to native speakers.\n"
            "- Contractions and informal expressions are allowed.\n"
            "- Preserve the original meaning.\n"
            "- Do not add or remove information.\n"
            "- Keep names, numbers, and formatting unchanged.\n"
            "- Output ONLY the translated text."
        ),
    },
    {
        "key": "game_formal",
        "label": "Game Translator — Formal",
        "prompt": (
            "Translate the game text using a formal tone suitable for game UI, menus, and system messages.\n"
            "\n"
            "CRITICAL RULES:\n"
            "- Preserve ALL placeholders EXACTLY as they appear.\n"
            "- NEVER translate, modify, reorder, or remove placeholders.\n"
            "- Preserve formatting and line breaks.\n"
            "\n"
            "Examples of placeholders (not limited to):\n"
            "{0} {1} {2} {name} {item}\n"
            "%s %d %i %f %u %l %g\n"
            "<tag> </tag>\n"
            "[tag]\n"
            "\\n \\t \\r \\\\n \\\\t\n"
            "\n"
            "Requirements:\n"
            "- Keep the meaning accurate.\n"
            "- Use concise formal language appropriate for UI text.\n"
            "- Output ONLY the translated text."
        ),
    },
    {
        "key": "game_casual",
        "label": "Game Translator — Casual",
        "prompt": (
            "Translate the game text in a casual and engaging tone suitable for character dialogue.\n"
            "\n"
            "CRITICAL RULES:\n"
            "- Preserve ALL placeholders EXACTLY as they appear.\n"
            "- NEVER translate, modify, reorder, or remove placeholders.\n"
            "- Preserve formatting and line breaks.\n"
            "\n"
            "Examples of placeholders (not limited to):\n"
            "{0} {1} {2} {name} {item}\n"
            "%s %d %i %f %u %l %g\n"
            "<tag> </tag>\n"
            "[tag]\n"
            "\\n \\t \\r \\\\n \\\\t\n"
            "\n"
            "Requirements:\n"
            "- Keep the original meaning.\n"
            "- Use lively and natural dialogue.\n"
            "- Output ONLY the translated text."
        ),
    },
    {
        "key": "technical",
        "label": "Technical / IT",
        "prompt": (
            "Translate technical or IT content accurately.\n"
            "Requirements:\n"
            "- Preserve technical terms, acronyms, and product names.\n"
            "- Do not translate code snippets, commands, or file paths.\n"
            "- Use industry-standard terminology.\n"
            "- Maintain formatting, punctuation, and capitalization when relevant.\n"
            "- Do not add explanations.\n"
            "- Output ONLY the translated text."
        ),
    },
    {
        "key": "subtitle",
        "label": "Subtitle / Dialogue",
        "prompt": (
            "Translate the text as subtitle dialogue.\n"
            "Requirements:\n"
            "- Keep sentences concise and easy to read.\n"
            "- Ensure the translation sounds natural when spoken.\n"
            "- Preserve the emotional tone of the original.\n"
            "- Avoid overly long sentences.\n"
            "- Output ONLY the translated subtitle text."
        ),
    },
    {
        "key": "ecommerce",
        "label": "E-Commerce / Product",
        "prompt": (
            "Translate the product description for e-commerce.\n"
            "Requirements:\n"
            "- Make the text persuasive and appealing to customers.\n"
            "- Preserve product names, specifications, measurements, and numbers.\n"
            "- Maintain clarity and readability.\n"
            "- Do not invent product features not present in the original.\n"
            "- Output ONLY the translated text."
        ),
    },
    {
        "key": "custom",
        "label": "Custom Prompt",
        "prompt": None,
    },
]


# ══════════════════════════════════════════════════════════════════════════════
#  PROVIDER CONFIGS
# ══════════════════════════════════════════════════════════════════════════════
PROVIDERS = [
    {
        "key":          "gemini",
        "label":        "Google Gemini",
        "url_models":   "https://generativelanguage.googleapis.com/v1beta/models",
        "url_generate": "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent",
        "api_key_hint": "AIza...",
    },
    {
        "key":          "groq",
        "label":        "Groq",
        "url_models":   "https://api.groq.com/openai/v1/models",
        "url_generate": "https://api.groq.com/openai/v1/chat/completions",
        "api_key_hint": "gsk_...",
    },
    {
        "key":          "cerebras",
        "label":        "Cerebras",
        "url_models":   "https://api.cerebras.ai/v1/models",
        "url_generate": "https://api.cerebras.ai/v1/chat/completions",
        "api_key_hint": "csk-...",
    },
]


# ══════════════════════════════════════════════════════════════════════════════
#  UI  &  LOG HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def _tw():
    return min(shutil.get_terminal_size((80, 24)).columns, 88)


def _div(char="─", color=GY):
    print(f"{color}{char * _tw()}{R}")


def print_banner():
    w = _tw()
    print(f"\n{GY}{'═' * w}{R}")
    print(f"{WH}{B}  Universal Excel Translator  v2.0{R}")
    print(f"{GY}  Powered by Gemini · Groq · Cerebras{R}")
    print(f"{GY}{'═' * w}{R}\n")


def print_step(title, color=CY):
    w = _tw()
    print(f"\n{color}{'─' * w}{R}")
    print(f"{color}{B}  {title}{R}")
    print(f"{color}{'─' * w}{R}\n")


# Structured log functions — professional developer style
def log_info(msg):   print(f"  {CY}[INFO]{R}   {GY}{msg}{R}")
def log_ok(msg):     print(f"  {GR}[OK]{R}     {WH}{msg}{R}")
def log_warn(msg):   print(f"  {YL}[WARN]{R}   {YL}{msg}{R}")
def log_err(msg):    print(f"  {RD}[ERROR]{R}  {RD}{msg}{R}")
def log_proc(msg):   print(f"  {BL}[PROC]{R}   {msg}")
def log_write(msg):  print(f"  {GY}[WRITE]{R}  {GY}{msg}{R}")
def log_repair(msg): print(f"  {MG}[REPAIR]{R} {msg}")
def log_fail(msg):   print(f"  {RD}[FAIL]{R}   {RD}{msg}{R}")


def print_menu(options, title=None, color=WH):
    if title:
        print(f"  {B}{YL}{title}{R}")
    for i, opt in enumerate(options, 1):
        print(f"  {GY}[{i}]{R}  {color}{opt}{R}")
    print()


def progress_bar(current, total, width=36, label=""):
    pct   = current / total if total > 0 else 0
    filled = int(width * pct)
    bar   = f"{GR}{'█' * filled}{GY}{'░' * (width - filled)}{R}"
    extra = f"  {GY}{label}{R}" if label else ""
    return f"  [{bar}] {WH}{current}/{total}{R} {GY}({pct:.0%}){R}{extra}"


def truncate(text, n=42):
    text = str(text).replace("\n", "↵").replace("\t", "→")
    return text if len(text) <= n else text[:n - 1] + "…"


def mask_key(key):
    key = key.strip()
    if len(key) > 8:
        return f"{key[:4]}{'*' * (len(key) - 8)}{key[-4:]}"
    return "****"


def _confirm(prompt="Continue? [Y/n]: "):
    try:
        ans = input(f"  {YL}{prompt}{R}").strip().lower()
        return ans in ("", "y", "yes")
    except (EOFError, KeyboardInterrupt):
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  NAVIGATION INPUT
# ══════════════════════════════════════════════════════════════════════════════
def nav_input(prompt, allow_empty=False):
    """
    Standard input with back/cancel handling.
    API keys: use this (visible while typing).
    Returns: stripped string, "BACK", or "CANCEL".
    """
    try:
        val = input(f"  {CY}{prompt}{R}  ")
    except (EOFError, KeyboardInterrupt):
        return "CANCEL"

    stripped = val.strip()
    lower    = stripped.lower()

    if lower in CANCEL_CMDS:
        print()
        if _confirm("Confirm cancel — all progress will be lost. [y/N]: "):
            print(f"\n  {YL}Cancelled. Goodbye.{R}\n")
            sys.exit(0)
        return nav_input(prompt, allow_empty)

    if lower in NAV_CMDS:
        return "BACK"

    if not allow_empty and stripped == "":
        log_err("Input cannot be empty.")
        return nav_input(prompt, allow_empty)

    return stripped


def nav_choose(options, prompt="Select: ", allow_back=True):
    """Choose from numbered list. Returns 0-based index, 'BACK', or 'CANCEL'."""
    while True:
        raw = nav_input(prompt)
        if raw == "BACK":
            if allow_back:
                return "BACK"
            log_err("Cannot go back at this step.")
            continue
        if raw == "CANCEL":
            return "CANCEL"
        try:
            idx = int(raw) - 1
            if 0 <= idx < len(options):
                return idx
        except ValueError:
            pass
        log_err(f"Enter a number between 1 and {len(options)}.")


# ══════════════════════════════════════════════════════════════════════════════
#  PARSE UTILITIES
# ══════════════════════════════════════════════════════════════════════════════
def parse_time_str(s):
    """'59s' | '1m30s' | '1m30.5s' | '90'  →  seconds (int)."""
    s = str(s).strip()
    m = re.match(r'^(?:(\d+)m)?(?:([\d.]+)s)?$', s)
    if m and (m.group(1) or m.group(2)):
        return max(1, int(int(m.group(1) or 0) * 60 + float(m.group(2) or 0)))
    try:
        return max(1, int(float(s)))
    except ValueError:
        return 5


def parse_retry_delay(response_text, headers):
    """
    Extract smart retry delay from API error response/headers.
    Priority: Gemini RetryInfo > Retry-After header > ratelimit-reset headers > message string.
    """
    delay = 5

    # 1. Gemini: error.details[].@type == RetryInfo → retryDelay
    try:
        data = json.loads(response_text)
        for d in data.get("error", {}).get("details", []):
            if "RetryInfo" in d.get("@type", "") and "retryDelay" in d:
                return max(delay, parse_time_str(d["retryDelay"]))
    except Exception:
        pass

    # 2. Standard Retry-After header (Groq, Cerebras)
    ra = (headers.get("retry-after") or headers.get("Retry-After", "")).strip()
    if ra:
        try:
            delay = max(delay, int(float(ra)))
        except ValueError:
            pass

    # 3. x-ratelimit-reset-requests  (e.g. "1m30.5s")
    for hdr in ("x-ratelimit-reset-requests", "x-ratelimit-reset-tokens",
                "x-ratelimit-reset-requests-day", "x-ratelimit-reset-tokens-minute"):
        val = (headers.get(hdr) or "").strip()
        if val:
            delay = max(delay, parse_time_str(val) + 1)

    # 4. Groq message fallback: "...try again in 46.72s..."
    try:
        msg = json.loads(response_text).get("error", {}).get("message", "")
        m = re.search(r'try again in ([\d.]+)s', msg, re.IGNORECASE)
        if m:
            delay = max(delay, int(float(m.group(1))) + 1)
    except Exception:
        pass

    return delay


def parse_cell_ref(ref):
    """'C2' → (col_index: int, row: int)  or  (None, None)."""
    m = re.match(r'^([A-Za-z]{1,3})(\d+)$', ref.strip())
    if not m:
        return None, None
    try:
        col = column_index_from_string(m.group(1).upper())
        row = int(m.group(2))
        return (col, row) if row >= 1 else (None, None)
    except Exception:
        return None, None


# ══════════════════════════════════════════════════════════════════════════════
#  KEY DISTRIBUTION ALGORITHMS
# ══════════════════════════════════════════════════════════════════════════════

class RoundRobinRotator:
    """Simple round-robin with per-key cooldown on 429."""

    name = "Round Robin"

    def __init__(self, keys):
        self.keys    = list(keys)
        self.current = 0
        self._blocked = {}   # key → unblock_timestamp

    def get_key(self):
        """Returns (key, wait_seconds). wait_seconds == 0 means ready."""
        now = time.time()
        for i in range(len(self.keys)):
            idx = (self.current + i) % len(self.keys)
            key = self.keys[idx]
            if self._blocked.get(key, 0) <= now:
                self.current = (idx + 1) % len(self.keys)
                return key, 0
        earliest = min(self.keys, key=lambda k: self._blocked.get(k, 0))
        return earliest, max(0, self._blocked[earliest] - now)

    def block(self, key, delay):
        self._blocked[key] = time.time() + delay

    def record_result(self, key, success, elapsed):
        pass   # not used by this algorithm

    def all_keys(self):
        return self.keys


class WeightedRotator:
    """
    Token-bucket weighted rotator.
    Keys with higher weight get proportionally more batches.
    Tokens refill every 60 s.
    """

    name = "Weighted Queue"
    REFILL_INTERVAL = 60

    def __init__(self, keys, weights):
        self.keys    = list(keys)
        self.weights = list(weights)
        self.tokens  = list(weights)
        self._blocked  = {}
        self._last_refill = time.time()

    def _refill(self):
        if time.time() - self._last_refill >= self.REFILL_INTERVAL:
            self.tokens = list(self.weights)
            self._last_refill = time.time()

    def get_key(self):
        self._refill()
        now = time.time()
        best_key, best_tokens, best_idx = None, -1, -1
        for i, key in enumerate(self.keys):
            if self._blocked.get(key, 0) > now:
                continue
            if self.tokens[i] > best_tokens:
                best_tokens, best_key, best_idx = self.tokens[i], key, i
        if best_key:
            if best_tokens > 0:
                self.tokens[best_idx] -= 1
            return best_key, 0
        earliest = min(self.keys, key=lambda k: self._blocked.get(k, 0))
        return earliest, max(0, self._blocked[earliest] - now)

    def block(self, key, delay):
        self._blocked[key] = time.time() + delay

    def record_result(self, key, success, elapsed):
        pass

    def all_keys(self):
        return self.keys


class AdaptiveRotator:
    """
    Self-optimising rotator.
    Scores each key by success rate, average response time, and idle time.
    No manual configuration needed.
    """

    name = "Adaptive Dynamic Queue"

    def __init__(self, keys):
        self.keys     = list(keys)
        self._blocked = {}
        self._stats   = {
            k: {"ok": 0, "fail": 0, "total_t": 0.0, "calls": 0, "last": 0.0}
            for k in keys
        }

    def _score(self, key):
        s   = self._stats[key]
        total = s["ok"] + s["fail"]
        sr  = s["ok"] / total if total > 0 else 0.5          # success rate 0-1
        avg = s["total_t"] / s["calls"] if s["calls"] > 0 else 1.0
        idle = min((time.time() - s["last"]) / 10.0, 3.0)    # max +3 bonus
        return sr * 5.0 + idle - avg * 0.3

    def get_key(self):
        now = time.time()
        candidates = [
            (self._score(k), k)
            for k in self.keys
            if self._blocked.get(k, 0) <= now
        ]
        if candidates:
            candidates.sort(reverse=True)
            return candidates[0][1], 0
        earliest = min(self.keys, key=lambda k: self._blocked.get(k, 0))
        return earliest, max(0, self._blocked[earliest] - now)

    def block(self, key, delay):
        self._blocked[key] = time.time() + delay

    def record_result(self, key, success, elapsed):
        s = self._stats[key]
        if success:
            s["ok"] += 1
        else:
            s["fail"] += 1
        s["total_t"] += elapsed
        s["calls"]   += 1
        s["last"]     = time.time()

    def all_keys(self):
        return self.keys


def build_rotator(keys, algorithm, weights=None):
    if algorithm == "weighted":
        w = weights if weights else [5] * len(keys)
        return WeightedRotator(keys, w)
    elif algorithm == "adaptive":
        return AdaptiveRotator(keys)
    else:
        return RoundRobinRotator(keys)


# ══════════════════════════════════════════════════════════════════════════════
#  PROVIDER API LAYER
# ══════════════════════════════════════════════════════════════════════════════

class ProviderError(Exception):
    def __init__(self, msg, status_code=0, is_rate_limit=False,
                 retry_delay=5, response_text="", headers=None):
        super().__init__(msg)
        self.status_code   = status_code
        self.is_rate_limit = is_rate_limit
        self.retry_delay   = retry_delay
        self.response_text = response_text
        self.headers       = headers or {}


def _openai_list_models(url, api_key):
    resp = requests.get(
        url, headers={"Authorization": f"Bearer {api_key}"}, timeout=15
    )
    if resp.status_code != 200:
        raise ProviderError(f"HTTP {resp.status_code}: {resp.text[:200]}",
                            status_code=resp.status_code)
    return sorted(m["id"] for m in resp.json().get("data", [])
                  if isinstance(m.get("id"), str))


def _openai_chat(url, api_key, model, messages, temperature=0.1):
    resp = requests.post(
        url,
        headers={"Authorization": f"Bearer {api_key}",
                 "Content-Type": "application/json"},
        json={"model": model, "messages": messages, "temperature": temperature},
        timeout=120,
    )
    hdrs, text = dict(resp.headers), resp.text
    if resp.status_code == 429:
        delay = parse_retry_delay(text, hdrs)
        raise ProviderError(f"Rate limit (429) — retry in {delay}s",
                            status_code=429, is_rate_limit=True,
                            retry_delay=delay, response_text=text, headers=hdrs)
    if resp.status_code != 200:
        raise ProviderError(f"HTTP {resp.status_code}: {text[:300]}",
                            status_code=resp.status_code,
                            response_text=text, headers=hdrs)
    return resp.json()["choices"][0]["message"]["content"]


def _gemini_list_models(api_key):
    url  = PROVIDERS[0]["url_models"]
    resp = requests.get(url, params={"key": api_key}, timeout=15)
    if resp.status_code != 200:
        raise ProviderError(f"HTTP {resp.status_code}: {resp.text[:200]}",
                            status_code=resp.status_code)
    models = []
    for m in resp.json().get("models", []):
        name    = m.get("name", "")
        methods = m.get("supportedGenerationMethods", [])
        if "generateContent" in methods and name.startswith("models/"):
            models.append(name[len("models/"):])
    return sorted(models)


def _gemini_generate(api_key, model, prompt_text, temperature=0.1):
    url  = PROVIDERS[0]["url_generate"].format(model=model)
    resp = requests.post(
        url,
        headers={"x-goog-api-key": api_key, "Content-Type": "application/json"},
        json={
            "contents": [{"parts": [{"text": prompt_text}]}],
            "generationConfig": {"temperature": temperature},
        },
        timeout=120,
    )
    hdrs, text = dict(resp.headers), resp.text
    if resp.status_code == 429:
        delay = parse_retry_delay(text, hdrs)
        raise ProviderError(f"Rate limit (429) — retry in {delay}s",
                            status_code=429, is_rate_limit=True,
                            retry_delay=delay, response_text=text, headers=hdrs)
    if resp.status_code != 200:
        raise ProviderError(f"HTTP {resp.status_code}: {text[:300]}",
                            status_code=resp.status_code,
                            response_text=text, headers=hdrs)
    try:
        return resp.json()["candidates"][0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError):
        raise ProviderError(f"Unexpected Gemini response: {text[:300]}")


def fetch_models(provider_key, api_key):
    if provider_key == "gemini":
        return _gemini_list_models(api_key)
    elif provider_key == "groq":
        return _openai_list_models(PROVIDERS[1]["url_models"], api_key)
    elif provider_key == "cerebras":
        return _openai_list_models(PROVIDERS[2]["url_models"], api_key)
    raise ValueError(f"Unknown provider: {provider_key}")


def call_llm(provider_key, api_key, model, prompt_text, temperature=0.1):
    if provider_key == "gemini":
        return _gemini_generate(api_key, model, prompt_text, temperature=temperature)
    elif provider_key == "groq":
        return _openai_chat(
            PROVIDERS[1]["url_generate"], api_key, model,
            [{"role": "user", "content": prompt_text}], temperature=temperature)
    elif provider_key == "cerebras":
        return _openai_chat(
            PROVIDERS[2]["url_generate"], api_key, model,
            [{"role": "user", "content": prompt_text}], temperature=temperature)
    raise ValueError(f"Unknown provider: {provider_key}")


# ══════════════════════════════════════════════════════════════════════════════
#  PROMPT BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

_TRANSLATE_SYS = (
    "You are a professional translator. "
    "Return ONLY the translated numbered lines. "
    "Do NOT add explanations, preamble, or extra commentary."
)

_REPAIR_SYS = (
    "You are a professional translation editor. "
    "Your ONLY job is to correct the TRANSLATED text — never the original. "
    "OUTPUT LANGUAGE must exactly match the Translated text language. "
    "CRITICAL: Do NOT translate the text into another language. "
    "CRITICAL: Do NOT output the source/original language. "
    "CRITICAL: If the Translated text is already correct, return it unchanged. "
    "Return ONLY the corrected numbered lines with [N] prefix. "
    "Do NOT explain, do NOT comment."
)


def build_translate_prompt(rows, src_lang, tgt_lang, style_prompt,
                            context_rows=None, global_context=None,
                            custom_placeholders=""):
    n       = len(rows)
    numbered = "\n".join(f"[{rn}] {txt}" for rn, txt in rows)

    ctx_block = ""
    if global_context and global_context.strip():
        ctx_block += f"\nGlobal context (applies to all lines):\n{global_context.strip()}\n"
    if context_rows:
        ctx_lines = "\n".join(f"[{rn}] {ctx}" for rn, ctx in context_rows)
        ctx_block += f"\nPer-row context (same row numbers):\n{ctx_lines}\n"

    ph_note = ""
    if custom_placeholders.strip():
        ph_note = f"\nAdditional placeholders to preserve exactly: {custom_placeholders.strip()}"

    return (
        f"{_TRANSLATE_SYS}\n\n"
        f"Translate {n} line(s) from {src_lang} to {tgt_lang}.\n"
        f"Style: {style_prompt}\n"
        f"{ctx_block}{ph_note}\n\n"
        f"RULES:\n"
        f"- Keep the [N] prefix in output\n"
        f"- Translate EACH line separately — do NOT merge or split\n"
        f"- Return EXACTLY {n} numbered line(s)\n"
        f"- Preserve all placeholder patterns exactly\n\n"
        f"Input:\n{numbered}\n\nOutput:"
    )


def build_translate_solo_prompt(text, src_lang, tgt_lang, style_prompt,
                                 context=None, global_context=None,
                                 custom_placeholders=""):
    ctx = ""
    if global_context and global_context.strip():
        ctx += f"Global context: {global_context.strip()}\n"
    if context and context.strip():
        ctx += f"Row context: {context.strip()}\n"
    ph_note = ""
    if custom_placeholders.strip():
        ph_note = f"Preserve these placeholders exactly: {custom_placeholders.strip()}\n"
    return (
        f"You are a professional translator.\n"
        f"Translate the following text from {src_lang} to {tgt_lang}.\n"
        f"Style: {style_prompt}\n"
        f"{ctx}{ph_note}\n"
        f"Return ONLY the translated text — nothing else.\n\n"
        f"Text: {text}\n\nTranslation:"
    )


def build_repair_prompt(rows_data, repair_mode, src_lang="", tgt_lang=""):
    n         = len(rows_data)
    row_nums  = [d['row_num'] for d in rows_data]
    example_n = row_nums[0]
    lang_note = (f"Translation language: {tgt_lang}. "
                 f"Source language: {src_lang}. "
                 if tgt_lang else "")

    if repair_mode == "repair":
        instruction = (
            f"{lang_note}"
            f"Review each TRANSLATED line for typos, grammar errors, and unnatural phrasing. "
            f"Output must be in {tgt_lang or 'the same language as the translation'}. "
            f"Do NOT translate back to the source language."
        )
        lines = [f"[{d['row_num']}] {d['translated']}" for d in rows_data]

    elif repair_mode == "repair_compare":
        instruction = (
            f"{lang_note}"
            f"The 'Original' is the SOURCE text. The 'Translated' is the TARGET text in {tgt_lang or 'target language'}. "
            f"Compare them for meaning accuracy, then fix typos and unnatural phrasing in the TRANSLATED text. "
            f"Your output MUST be in {tgt_lang or 'the translation language'} — never in the source language."
        )
        lines = [
            f"[{d['row_num']}]\n  Original   : {d.get('original','')}\n"
            f"  Translated : {d['translated']}"
            for d in rows_data
        ]

    else:  # repair_compare_context
        instruction = (
            f"{lang_note}"
            f"The 'Original' is SOURCE text. The 'Translated' is the TARGET text in {tgt_lang or 'target language'}. "
            f"Use the Context and Original to verify meaning accuracy, then fix any errors in the TRANSLATED text. "
            f"Your output MUST be in {tgt_lang or 'the translation language'} — never in the source language."
        )
        lines = [
            f"[{d['row_num']}]\n  Context    : {d.get('context','')}\n"
            f"  Original   : {d.get('original','')}\n"
            f"  Translated : {d['translated']}"
            for d in rows_data
        ]

    numbered = "\n".join(lines)
    nums_str  = ", ".join(str(r) for r in row_nums)
    return (
        f"{_REPAIR_SYS}\n\n"
        f"{instruction}\n\n"
        f"CRITICAL OUTPUT FORMAT:\n"
        f"- You MUST output EXACTLY {n} line(s) with [N] prefix\n"
        f"- Row numbers to output (in order): {nums_str}\n"
        f"- Format each line EXACTLY as: [ROW_NUMBER] corrected text here\n"
        f"- Example: [{example_n}] Corrected translation in {tgt_lang or 'target language'} here.\n"
        f"- Do NOT skip the [N] prefix — it is required for parsing\n"
        f"- Do NOT output the Original — output ONLY the corrected Translated text\n"
        f"- If a translation is already correct, return it unchanged with the [N] prefix\n"
        f"- Preserve all placeholder patterns exactly (e.g. {{0}}, %s, \\n)\n\n"
        f"Input:\n{numbered}\n\nOutput:"
    )


def build_repair_solo_prompt(translated, repair_mode, original=None, context=None,
                              src_lang="", tgt_lang=""):
    lang_note = (f"Output language: {tgt_lang}. Do NOT output in {src_lang}.\n"
                 if tgt_lang else "")

    if repair_mode == "repair":
        instruction = (
            f"{lang_note}"
            f"Fix typos, grammar errors, and unnatural phrasing in this translation. "
            f"Output must be in {tgt_lang or 'the same language as the translation'}."
        )
        body = f"Translation: {translated}"
    elif repair_mode == "repair_compare":
        instruction = (
            f"{lang_note}"
            f"The Original is the SOURCE. The Translation is in {tgt_lang or 'target language'}. "
            f"Fix inaccuracies and unnatural phrasing in the Translation. "
            f"Output ONLY the corrected Translation in {tgt_lang or 'target language'}."
        )
        body = f"Original   : {original}\nTranslation: {translated}"
    else:
        instruction = (
            f"{lang_note}"
            f"The Original is the SOURCE. The Translation is in {tgt_lang or 'target language'}. "
            f"Use Context and Original to verify meaning, then fix the Translation. "
            f"Output ONLY the corrected Translation in {tgt_lang or 'target language'}."
        )
        body = f"Context    : {context}\nOriginal   : {original}\nTranslation: {translated}"
    return (
        f"You are a professional translation editor.\n"
        f"{instruction}\n\n"
        f"{body}\n\n"
        f"Return ONLY the corrected translation — nothing else.\n\nCorrected:"
    )


# ══════════════════════════════════════════════════════════════════════════════
#  RESPONSE PARSER
# ══════════════════════════════════════════════════════════════════════════════

def parse_batch_response(text, expected_row_nums):
    """
    Parse [N] lines from LLM output.
    Strategy 1: strict  [N] prefix regex
    Strategy 2: loose   N. or N) prefix
    Strategy 3: line-by-line positional fallback
    Returns dict: row_num → translated_str
    """
    def _clean(val):
        """Strip any leaked [N], [N][M], or literal [N] prefixes from value."""
        val = val.strip()
        # Strip patterns like [N], [760], [N][760], [760][N] from start
        val = re.sub(r'^(\[[\dN]+\]\s*)+', '', val).strip()
        # Strip patterns like N. or N) from start
        val = re.sub(r'^\d+[.)]\s+', '', val).strip()
        return val

    result = {}

    # Strategy 1: strict [N] prefix  e.g.  [2] text
    for m in re.finditer(r'\[(\d+)\]\s*(.*?)(?=\n\s*\[\d+\]|\Z)', text, re.DOTALL):
        rn  = int(m.group(1))
        val = _clean(m.group(2))
        if val:
            result[rn] = val

    if len(result) >= len(expected_row_nums) * 0.8:
        return result

    # Strategy 2: loose prefix  e.g.  2. text  or  2) text
    result2 = {}
    for m in re.finditer(r'^\s*(\d+)[.)]\s+(.*)', text, re.MULTILINE):
        rn  = int(m.group(1))
        val = _clean(m.group(2))
        if rn in expected_row_nums and val:
            result2[rn] = val

    if len(result2) >= len(expected_row_nums) * 0.8:
        return result2

    # Merge partial results from strategy 1 + 2
    merged = {**result2, **result}
    if len(merged) >= len(expected_row_nums) * 0.8:
        return merged

    # Strategy 3: positional line-by-line fallback
    lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
    result3 = {}
    for i, rn in enumerate(expected_row_nums):
        if i < len(lines):
            result3[rn] = _clean(lines[i])
    return result3


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL HANDLER
# ══════════════════════════════════════════════════════════════════════════════

def scan_excel_files(directory="."):
    files = sorted(
        glob.glob(os.path.join(directory, "*.xlsx")) +
        glob.glob(os.path.join(directory, "*.xls"))
    )
    return [os.path.basename(f) for f in files]


def load_column(filepath, col_idx, start_row, num_rows):
    """Return list of (row_num, text_or_empty)."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = []
    for r in range(start_row, start_row + num_rows):
        val  = ws.cell(row=r, column=col_idx).value
        rows.append((r, str(val).strip() if val is not None else ""))
    wb.close()
    return rows


def count_filled_rows(filepath, col_idx, start_row, num_rows):
    """Count rows in dst column that already have data."""
    wb  = load_workbook(filepath, data_only=True)
    ws  = wb.active
    count = 0
    for r in range(start_row, start_row + num_rows):
        val = ws.cell(row=r, column=col_idx).value
        if val is not None and str(val).strip():
            count += 1
    wb.close()
    return count


def write_batch(filepath, col_idx, results):
    """Write {row_num: text} to file immediately after each batch."""
    wb = load_workbook(filepath)
    ws = wb.active
    for row_num, text in results.items():
        ws.cell(row=row_num, column=col_idx, value=text)
    wb.save(filepath)
    wb.close()


def preview_rows(filepath, col_indices_labels, start_row, n=3):
    wb   = load_workbook(filepath, data_only=True)
    ws   = wb.active
    rows = []
    for r in range(start_row, start_row + n):
        row_data = {"_row": r}
        for col_idx, label in col_indices_labels:
            val = ws.cell(row=r, column=col_idx).value
            row_data[label] = str(val).strip() if val is not None else "(empty)"
        rows.append(row_data)
    wb.close()
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  PROCESSING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

FAIL_MARKER      = "[TRANSLATE_FAILED]"
MAX_RETRIES      = 3          # default, overridden by job config
RETRY_DELAY_503  = 45         # seconds to wait on 503
RETRY_DELAY_ERR  = 5          # seconds to wait on other errors


class ProcessingEngine:
    """
    Unified engine for Translate, Translate+Context, Translate+Repair,
    and all Repair Mode variants.
    Writes results to Excel after EVERY batch (crash-safe).

    repair_model: if set, used for repair calls instead of self.model
    """

    def __init__(self, provider_key, rotator, model, job, repair_model=None):
        self.provider     = provider_key
        self.rotator      = rotator
        self.model        = model
        self.repair_model = repair_model if repair_model else model
        self.job          = job
        self.stats        = {"ok": 0, "repaired": 0, "skipped": 0, "failed": 0, "total": 0}
        self._write_lock  = threading.Lock()
        self._stats_lock  = threading.Lock()
        self._print_lock  = threading.Lock()

    # ── Public entry points ───────────────────────────────────────────────────

    def run_translate(self, source_rows, context_map=None):
        job         = self.job
        out_path    = job["out_path"]
        dst_col     = job["translated_col"]
        with_repair = job.get("with_repair", False)
        repair_col  = job.get("repair_col")
        skip_filled = job.get("skip_filled", True)

        to_process = [(rn, txt) for rn, txt in source_rows if txt.strip()]
        skip_rows  = [(rn, txt) for rn, txt in source_rows if not txt.strip()]

        self.stats["total"]   = len(source_rows)
        self.stats["skipped"] = len(skip_rows)

        if skip_filled:
            to_process = self._filter_resume(out_path, dst_col, to_process)

        batches = self._make_batches(to_process, job["batch_size"])
        total_b = len(batches)

        # ── Phase 1: Translate ────────────────────────────────────────────────
        all_translated = {}   # rn → translated text (used by repair phase)

        for b_idx, batch in enumerate(batches):
            log_proc(
                f"Batch {b_idx+1}/{total_b}  "
                f"rows {batch[0][0]}-{batch[-1][0]}  "
                f"({len(batch)} rows)"
            )

            translated = self._translate_batch(batch, context_map, b_idx + 1, total_b)

            if translated:
                write_batch(out_path, dst_col, translated)
                log_write(f"Translate → col {get_column_letter(dst_col)}  "
                          f"rows {batch[0][0]}-{batch[-1][0]}")
                all_translated.update(translated)

            done = self.stats["ok"] + self.stats["skipped"] + self.stats["failed"]
            print(progress_bar(done, self.stats["total"],
                               label=f"ok:{self.stats['ok']} "
                                     f"fail:{self.stats['failed']} "
                                     f"skip:{self.stats['skipped']}"))

            if b_idx < total_b - 1 and job.get("inter_delay", 0) > 0:
                time.sleep(job["inter_delay"])

        # ── Phase 2: Repair (same algorithm, after all translate done) ────────
        if with_repair and repair_col and all_translated:
            self._run_repair_phase(source_rows, all_translated, context_map)

    def run_repair(self, rows_to_repair):
        job        = self.job
        out_path   = job["out_path"]
        repair_col = job["repair_col"]
        skip_filled = job.get("skip_filled", True)
        repair_mode = job["repair_mode"]

        to_process = [d for d in rows_to_repair if d.get("translated", "").strip()]
        skipped    = [d for d in rows_to_repair if not d.get("translated", "").strip()]

        self.stats["total"]   = len(rows_to_repair)
        self.stats["skipped"] = len(skipped)

        if skip_filled:
            filled_rows = self._get_filled_rows(out_path, repair_col,
                                                job["start_row"], job["num_rows"])
            to_process = [d for d in to_process if d["row_num"] not in filled_rows]
            self.stats["skipped"] += len(rows_to_repair) - len(skipped) - len(to_process)

        batches = self._make_batches_repair(to_process, job["batch_size"])
        total_b = len(batches)

        for b_idx, batch in enumerate(batches):
            log_repair(
                f"Batch {b_idx+1}/{total_b}  "
                f"rows {batch[0]['row_num']}-{batch[-1]['row_num']}"
            )

            repaired = self._repair_batch(batch, repair_mode, b_idx + 1, total_b)
            if repaired:
                write_batch(out_path, repair_col, repaired)
                log_write(f"Repair    → col {get_column_letter(repair_col)}  "
                          f"rows {batch[0]['row_num']}-{batch[-1]['row_num']}")

            done = self.stats["ok"] + self.stats["skipped"] + self.stats["failed"]
            print(progress_bar(done, self.stats["total"],
                               label=f"ok:{self.stats['ok']} "
                                     f"fail:{self.stats['failed']} "
                                     f"skip:{self.stats['skipped']}"))

            if b_idx < total_b - 1 and job.get("inter_delay", 0) > 0:
                time.sleep(job["inter_delay"])

    # ── Parallel entry points ─────────────────────────────────────────────────

    def run_translate_parallel(self, source_rows, context_map=None, n_workers=1):
        """
        Parallel translate: divide batches into waves of n_workers.
        Each worker in a wave uses a pinned API key.
        Crash-safe: writes to Excel after each batch completes.
        """
        job         = self.job
        out_path    = job["out_path"]
        dst_col     = job["translated_col"]
        repair_col  = job.get("repair_col")
        with_repair = job.get("with_repair", False)
        skip_filled = job.get("skip_filled", True)
        keys        = self.rotator.all_keys()
        n_workers   = min(n_workers, len(keys))

        to_process = [(rn, txt) for rn, txt in source_rows if txt.strip()]
        skip_rows  = [(rn, txt) for rn, txt in source_rows if not txt.strip()]

        self.stats["total"]   = len(source_rows)
        self.stats["skipped"] = len(skip_rows)

        if skip_filled:
            to_process = self._filter_resume(out_path, dst_col, to_process)

        batches = self._make_batches(to_process, job["batch_size"])
        total_b = len(batches)

        if total_b == 0:
            return

        log_info(f"Parallel mode  : {n_workers} workers  |  {total_b} batches  |  "
                 f"{len(keys)} keys")

        # Split batches into waves
        waves = [batches[i:i + n_workers] for i in range(0, total_b, n_workers)]
        batch_counter = [0]

        all_translated = {}   # rn → translated (collected across all waves)

        def process_one_translate(batch, key, wave_idx, slot_idx, total_b):
            """Single translate task — runs in its own thread. Returns (batch, result)."""
            b_num = wave_idx * n_workers + slot_idx + 1
            with self._print_lock:
                log_proc(f"[W{wave_idx+1}·S{slot_idx+1}] Batch {b_num}/{total_b}  "
                         f"rows {batch[0][0]}-{batch[-1][0]}  key ...{key[-4:]}")

            translated = self._translate_batch_keyed(
                batch, key, context_map, b_num, total_b
            )
            return batch, translated

        for w_idx, wave in enumerate(waves):
            wave_results = {}   # rn → translated  (collected from all threads in wave)

            with ThreadPoolExecutor(max_workers=n_workers) as executor:
                futures = {}
                for s_idx, batch in enumerate(wave):
                    key = keys[s_idx % len(keys)]
                    f   = executor.submit(
                        process_one_translate, batch, key, w_idx, s_idx, total_b
                    )
                    futures[f] = batch

                for f in as_completed(futures):
                    try:
                        batch, translated = f.result()
                        if translated:
                            wave_results.update(translated)
                    except Exception as e:
                        with self._print_lock:
                            log_err(f"Parallel batch error: {e}")

            # ── 1x write for entire wave ──────────────────────────────────────
            if wave_results:
                write_batch(out_path, dst_col, wave_results)
                first = min(wave_results)
                last  = max(wave_results)
                log_write(f"Translate → col {get_column_letter(dst_col)}  "
                          f"rows {first}-{last}  ({len(wave_results)} rows, wave {w_idx+1})")
                all_translated.update(wave_results)

            done = self.stats["ok"] + self.stats["skipped"] + self.stats["failed"]
            print(progress_bar(done, self.stats["total"],
                               label=f"ok:{self.stats['ok']} "
                                     f"fail:{self.stats['failed']} "
                                     f"skip:{self.stats['skipped']} "
                                     f"wave:{w_idx+1}/{len(waves)}"))

            if w_idx < len(waves) - 1 and job.get("inter_delay", 0) > 0:
                time.sleep(job["inter_delay"])

        # ── Phase 2: Repair — same algorithm & workers as translate ───────────
        if with_repair and repair_col and all_translated:
            self._run_repair_phase(source_rows, all_translated, context_map,
                                   n_workers=n_workers, parallel=True)

    def run_repair_parallel(self, rows_to_repair, n_workers=1):
        """
        Parallel repair: divide batches into waves of n_workers.
        """
        job         = self.job
        out_path    = job["out_path"]
        repair_col  = job["repair_col"]
        skip_filled = job.get("skip_filled", True)
        repair_mode = job["repair_mode"]
        keys        = self.rotator.all_keys()
        n_workers   = min(n_workers, len(keys))

        to_process = [d for d in rows_to_repair if d.get("translated", "").strip()]
        skipped    = [d for d in rows_to_repair if not d.get("translated", "").strip()]

        self.stats["total"]   = len(rows_to_repair)
        self.stats["skipped"] = len(skipped)

        if skip_filled:
            filled_rows = self._get_filled_rows(out_path, repair_col,
                                                job["start_row"], job["num_rows"])
            to_process = [d for d in to_process if d["row_num"] not in filled_rows]
            self.stats["skipped"] += len(rows_to_repair) - len(skipped) - len(to_process)

        batches = self._make_batches_repair(to_process, job["batch_size"])
        total_b = len(batches)

        if total_b == 0:
            return

        log_info(f"Parallel repair: {n_workers} workers  |  {total_b} batches  |  "
                 f"{len(keys)} keys")

        waves = [batches[i:i + n_workers] for i in range(0, total_b, n_workers)]

        def process_one_repair(batch, key, wave_idx, slot_idx, total_b):
            b_num = wave_idx * n_workers + slot_idx + 1
            with self._print_lock:
                log_repair(f"[W{wave_idx+1}·S{slot_idx+1}] Batch {b_num}/{total_b}  "
                           f"rows {batch[0]['row_num']}-{batch[-1]['row_num']}  "
                           f"key ...{key[-4:]}")

            repaired = self._repair_batch_keyed(batch, key, repair_mode, b_num, total_b)

            if repaired:
                with self._write_lock:
                    write_batch(out_path, repair_col, repaired)
                with self._print_lock:
                    log_write(f"Repair    → col {get_column_letter(repair_col)}  "
                              f"rows {batch[0]['row_num']}-{batch[-1]['row_num']}")

            return repaired

        for w_idx, wave in enumerate(waves):
            with ThreadPoolExecutor(max_workers=n_workers) as executor:
                futures = {}
                for s_idx, batch in enumerate(wave):
                    key = keys[s_idx % len(keys)]
                    f   = executor.submit(
                        process_one_repair, batch, key, w_idx, s_idx, total_b
                    )
                    futures[f] = batch

                for f in as_completed(futures):
                    try:
                        f.result()
                    except Exception as e:
                        with self._print_lock:
                            log_err(f"Parallel repair error: {e}")

            done = self.stats["ok"] + self.stats["skipped"] + self.stats["failed"]
            print(progress_bar(done, self.stats["total"],
                               label=f"ok:{self.stats['ok']} "
                                     f"fail:{self.stats['failed']} "
                                     f"skip:{self.stats['skipped']} "
                                     f"wave:{w_idx+1}/{len(waves)}"))

            if w_idx < len(waves) - 1 and job.get("inter_delay", 0) > 0:
                time.sleep(job["inter_delay"])

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _make_batches(self, rows, size):
        return [rows[i:i + size] for i in range(0, len(rows), size)]

    def _make_batches_repair(self, rows, size):
        return [rows[i:i + size] for i in range(0, len(rows), size)]

    def _filter_resume(self, filepath, col_idx, rows):
        filled = self._get_filled_rows(filepath, col_idx,
                                       self.job["start_row"], self.job["num_rows"])
        skipped = [(rn, txt) for rn, txt in rows if rn in filled]
        remaining = [(rn, txt) for rn, txt in rows if rn not in filled]
        if skipped:
            log_info(f"Resuming: {len(skipped)} row(s) already done, skipping.")
            self.stats["skipped"] += len(skipped)
        return remaining

    def _get_filled_rows(self, filepath, col_idx, start_row, num_rows):
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            filled = set()
            for r in range(start_row, start_row + num_rows):
                val = ws.cell(row=r, column=col_idx).value
                if val is not None and str(val).strip():
                    filled.add(r)
            wb.close()
            return filled
        except Exception:
            return set()

    # ── Translate batch ────────────────────────────────────────────────────────

    def _translate_batch(self, batch, context_map, b_idx, total_b):
        job       = self.job
        mode      = job.get("op_mode", "translate")
        src_lang  = job.get("src_lang", "")
        debug_log = job.get("debug_log", False)

        # ── Language pre-filter ───────────────────────────────────────────────
        send_batch, skip_map = [], {}
        if _LANGDETECT_OK and src_lang:
            for rn, txt in batch:
                try:
                    detected = _langdetect(txt)
                    src_code = src_lang.lower()[:2]
                    if detected != src_code and not txt.strip().isascii():
                        skip_map[rn] = txt   # copy as-is
                        continue
                except Exception:
                    pass
                send_batch.append((rn, txt))
        else:
            send_batch = list(batch)

        expected  = [rn for rn, _ in send_batch]
        ctx_rows  = None
        if mode == "translate_context" and context_map:
            ctx_rows = [(rn, context_map.get(rn, "")) for rn, _ in send_batch]

        result = dict(skip_map)  # pre-fill with skipped rows
        for rn, txt in skip_map.items():
            print(f"  {GY}Row {rn:>5}{R}  {GY}[skip — not {src_lang}]{R}  {C2}{truncate(txt, 50)}{R}")
            self.stats["skipped"] += 1

        if not send_batch:
            return result

        prompt = build_translate_prompt(
            send_batch,
            src_lang=src_lang,
            tgt_lang=job["tgt_lang"],
            style_prompt=job["style_prompt"],
            context_rows=ctx_rows,
            global_context=job.get("global_context"),
            custom_placeholders=job.get("custom_placeholders", ""),
        )

        raw_response = self._call_with_retry(prompt, f"Batch {b_idx}/{total_b}",
                                             use_repair_model=False)
        if raw_response is None:
            for rn, _ in send_batch:
                self.stats["failed"] += 1
            result.update({rn: FAIL_MARKER for rn, _ in send_batch})
            return result

        parsed = parse_batch_response(raw_response, expected)

        # ── Debug log if parse is partial ────────────────────────────────────
        missing = [rn for rn in expected if not parsed.get(rn, "").strip()]
        if missing and debug_log:
            log_warn(f"Batch {b_idx}: parse partial {len(parsed)}/{len(expected)}")
            print(f"  {GY}┌─ Raw AI response ({'first 800 chars'}):{'─'*20}{R}")
            for line in raw_response[:800].splitlines():
                print(f"  {GY}│{R} {line}")
            print(f"  {GY}└{'─'*50}{R}")

        # ── Mini-batch retry for missing rows ─────────────────────────────────
        if missing:
            log_warn(f"Batch {b_idx}: {len(missing)} row(s) missing → mini-batch retry")
            mini_batch = [(rn, txt) for rn, txt in send_batch if rn in missing]
            mini_prompt = build_translate_prompt(
                mini_batch,
                src_lang=src_lang,
                tgt_lang=job["tgt_lang"],
                style_prompt=job["style_prompt"],
                context_rows=[(rn, context_map.get(rn,"")) for rn, _ in mini_batch]
                              if context_map and mode == "translate_context" else None,
                global_context=job.get("global_context"),
                custom_placeholders=job.get("custom_placeholders", ""),
            )
            mini_raw = self._call_with_retry(mini_prompt, f"Mini-batch {b_idx}",
                                             use_repair_model=False)
            if mini_raw:
                mini_parsed = parse_batch_response(mini_raw, missing)
                parsed.update(mini_parsed)
                still_missing = [rn for rn in missing if not parsed.get(rn,"").strip()]
                if still_missing:
                    log_warn(f"Mini-batch: {len(still_missing)} still missing → solo retry")
                else:
                    log_ok(f"Mini-batch retry recovered {len(missing)} row(s)")

        for rn, src_text in send_batch:
            val = parsed.get(rn, "").strip()
            if val:
                result[rn] = val
                src_d = truncate(src_text, 36)
                res_d = truncate(val, 36)
                print(f"  {GY}Row {rn:>5}{R}  {C2}{src_d}{R}  {GY}→{R}  {GR}{res_d}{R}")
                self.stats["ok"] += 1
            else:
                solo = self._solo_translate(rn, src_text, context_map, b_idx, total_b)
                result[rn] = solo
                if solo == FAIL_MARKER:
                    self.stats["failed"] += 1
                else:
                    self.stats["ok"] += 1

        return result

    def _solo_translate(self, row_num, text, context_map, b_idx, total_b):
        job  = self.job
        mode = job.get("op_mode", "translate")
        ctx  = context_map.get(row_num, "") if context_map else None
        log_warn(f"Row {row_num}: parse error in batch → retrying solo")
        prompt = build_translate_solo_prompt(
            text,
            src_lang=job["src_lang"],
            tgt_lang=job["tgt_lang"],
            style_prompt=job["style_prompt"],
            context=ctx,
            global_context=job.get("global_context"),
            custom_placeholders=job.get("custom_placeholders", ""),
        )
        response = self._call_with_retry(prompt, f"Solo Row {row_num}",
                                         use_repair_model=False)
        if response:
            result = response.strip().split("\n")[0].strip()
            result = re.sub(r'^\[?\d+\]?\s*', '', result)
            log_ok(f"Row {row_num}: solo retry success  →  {truncate(result, 40)}")
            return result
        log_fail(f"Row {row_num}: solo retry failed → {FAIL_MARKER}")
        return FAIL_MARKER

    # ── Repair batch (after translate) ────────────────────────────────────────

    # ── Repair phase (called after translate phase completes) ────────────────

    def _run_repair_phase(self, source_rows, all_translated, context_map,
                          n_workers=1, parallel=False):
        """
        Run repair as a clean separate phase after translate is done.
        Uses same algorithm/workers as translate phase.
        """
        job        = self.job
        out_path   = job["out_path"]
        repair_col = job.get("repair_col")
        repair_mode = job.get("repair_mode") or (
            "repair_compare_context" if context_map else "repair_compare"
        )
        keys = self.rotator.all_keys()

        # Build rows_data from translate results
        src_map = {rn: txt for rn, txt in source_rows}
        rows_data = []
        for rn, translated in all_translated.items():
            if not translated or translated == FAIL_MARKER:
                continue
            d = {"row_num": rn, "translated": translated,
                 "original": src_map.get(rn, "")}
            if context_map:
                d["context"] = context_map.get(rn, "")
            rows_data.append(d)

        rows_data.sort(key=lambda x: x["row_num"])

        if not rows_data:
            return

        batches = self._make_batches_repair(rows_data, job["batch_size"])
        total_b = len(batches)

        _div()
        log_proc(f"{'Parallel ' if parallel else ''}Repair phase  |  "
                 f"{len(rows_data)} rows  |  {total_b} batches  |  "
                 f"{'%d workers' % n_workers if parallel else 'sequential'}")
        print()

        if parallel and n_workers > 1:
            # ── Parallel repair — same wave system as translate ───────────────
            waves = [batches[i:i + n_workers] for i in range(0, total_b, n_workers)]

            def process_one_repair(batch, key, wave_idx, slot_idx, total_b):
                b_num = wave_idx * n_workers + slot_idx + 1
                with self._print_lock:
                    log_repair(f"[W{wave_idx+1}·S{slot_idx+1}] Repair {b_num}/{total_b}  "
                               f"rows {batch[0]['row_num']}-{batch[-1]['row_num']}  "
                               f"key ...{key[-4:]}")
                repaired = self._repair_batch_keyed(batch, key, repair_mode, b_num, total_b)
                return batch, repaired

            for w_idx, wave in enumerate(waves):
                wave_repaired = {}

                with ThreadPoolExecutor(max_workers=n_workers) as executor:
                    futures = {}
                    for s_idx, batch in enumerate(wave):
                        key = keys[s_idx % len(keys)]
                        f   = executor.submit(
                            process_one_repair, batch, key, w_idx, s_idx, total_b
                        )
                        futures[f] = batch

                    for f in as_completed(futures):
                        try:
                            batch, repaired = f.result()
                            if repaired:
                                wave_repaired.update(repaired)
                        except Exception as e:
                            with self._print_lock:
                                log_err(f"Parallel repair error: {e}")

                # ── 1x write for entire wave ──────────────────────────────────
                if wave_repaired:
                    write_batch(out_path, repair_col, wave_repaired)
                    first = min(wave_repaired)
                    last  = max(wave_repaired)
                    log_write(f"Repair    → col {get_column_letter(repair_col)}  "
                              f"rows {first}-{last}  ({len(wave_repaired)} rows, wave {w_idx+1})")

                done = self.stats["repaired"] + self.stats["failed"]
                print(progress_bar(done, len(rows_data),
                                   label=f"repaired:{self.stats['repaired']} "
                                         f"fail:{self.stats['failed']} "
                                         f"wave:{w_idx+1}/{len(waves)}"))

                if w_idx < len(waves) - 1 and job.get("inter_delay", 0) > 0:
                    time.sleep(job["inter_delay"])

        else:
            # ── Sequential repair ─────────────────────────────────────────────
            for b_idx, batch in enumerate(batches):
                log_repair(f"Repair {b_idx+1}/{total_b}  "
                           f"rows {batch[0]['row_num']}-{batch[-1]['row_num']}")

                repaired = self._repair_batch(batch, repair_mode, b_idx + 1, total_b)
                if repaired:
                    write_batch(out_path, repair_col, repaired)
                    log_write(f"Repair    → col {get_column_letter(repair_col)}  "
                              f"rows {batch[0]['row_num']}-{batch[-1]['row_num']}")

                done = self.stats["repaired"] + self.stats["failed"]
                print(progress_bar(done, len(rows_data),
                                   label=f"repaired:{self.stats['repaired']} "
                                         f"fail:{self.stats['failed']}"))

                if b_idx < total_b - 1 and job.get("inter_delay", 0) > 0:
                    time.sleep(job["inter_delay"])

    def _repair_batch_after_translate(self, batch, translated, context_map):
        job         = self.job
        # Use explicitly configured repair_mode if set, else auto-detect from context
        configured  = job.get("repair_mode")
        if configured:
            repair_mode = configured
        else:
            repair_mode = "repair_compare_context" if context_map else "repair_compare"

        rows_data = []
        for rn, src_text in batch:
            t = translated.get(rn, "")
            if not t or t == FAIL_MARKER:
                continue
            d = {"row_num": rn, "translated": t, "original": src_text}
            if context_map:
                d["context"] = context_map.get(rn, "")
            rows_data.append(d)

        if not rows_data:
            return {}

        return self._repair_batch(rows_data, repair_mode, label="Repair after translate")

    # ── Repair batch (standalone) ─────────────────────────────────────────────

    def _repair_batch(self, rows_data, repair_mode, b_idx=None, total_b=None, label=None):
        expected  = [d["row_num"] for d in rows_data]
        lbl       = label or (f"Repair {b_idx}/{total_b}" if b_idx else "Repair")
        debug_log = self.job.get("debug_log", False)
        prompt    = build_repair_prompt(rows_data, repair_mode,
                                        src_lang=self.job.get("src_lang", ""),
                                        tgt_lang=self.job.get("tgt_lang", ""))

        raw = self._call_with_retry(prompt, lbl, use_repair_model=True)
        if raw is None:
            for d in rows_data:
                self.stats["failed"] += 1
            return {d["row_num"]: FAIL_MARKER for d in rows_data}

        parsed  = parse_batch_response(raw, expected)
        missing = [rn for rn in expected if not parsed.get(rn, "").strip()]

        if missing and debug_log:
            log_warn(f"Repair {b_idx}: parse partial {len(parsed)}/{len(expected)}")
            print(f"  {GY}┌─ Raw repair response:{'─'*30}{R}")
            for line in raw[:800].splitlines():
                print(f"  {GY}│{R} {line}")
            print(f"  {GY}└{'─'*50}{R}")

        if missing:
            log_warn(f"Repair {b_idx}: {len(missing)} row(s) missing → mini-batch retry")
            mini_data   = [d for d in rows_data if d["row_num"] in missing]
            mini_prompt = build_repair_prompt(mini_data, repair_mode,
                                              src_lang=self.job.get("src_lang",""),
                                              tgt_lang=self.job.get("tgt_lang",""))
            mini_raw = self._call_with_retry(mini_prompt, f"Mini-repair {b_idx}",
                                             use_repair_model=True)
            if mini_raw:
                mini_parsed = parse_batch_response(mini_raw, missing)
                parsed.update(mini_parsed)
                still = [rn for rn in missing if not parsed.get(rn,"").strip()]
                if not still:
                    log_ok(f"Mini-batch repair recovered {len(missing)} row(s)")

        result = {}
        for d in rows_data:
            rn  = d["row_num"]
            val = parsed.get(rn, "").strip()
            if val:
                orig_d   = truncate(d["translated"], 34)
                fixed_d  = truncate(val, 34)
                changed  = "(repaired)" if val != d["translated"] else "(no change)"
                print(f"  {GY}Row {rn:>5}{R}  {C2}{orig_d}{R}  {GY}→{R}  "
                      f"{MG}{fixed_d}{R}  {GY}{changed}{R}")
                result[rn] = val
                self.stats["repaired"] += 1
                self.stats["ok"]       += 1
            else:
                solo = self._solo_repair(d, repair_mode)
                result[rn] = solo
                if solo == FAIL_MARKER:
                    self.stats["failed"] += 1
                else:
                    self.stats["repaired"] += 1
                    self.stats["ok"]       += 1

        return result

    def _solo_repair(self, row_data, repair_mode):
        rn = row_data["row_num"]
        log_warn(f"Row {rn}: repair parse error → retrying solo")
        prompt = build_repair_solo_prompt(
            translated=row_data["translated"],
            repair_mode=repair_mode,
            original=row_data.get("original"),
            context=row_data.get("context"),
            src_lang=self.job.get("src_lang", ""),
            tgt_lang=self.job.get("tgt_lang", ""),
        )
        # Use repair model for solo repair too
        response = self._call_with_retry(prompt, f"Solo Repair {rn}",
                                         use_repair_model=True)
        if response:
            result = response.strip().split("\n")[0].strip()
            result = re.sub(r'^\[?\d+\]?\s*', '', result)
            log_ok(f"Row {rn}: solo repair success  →  {truncate(result, 40)}")
            return result
        log_fail(f"Row {rn}: solo repair failed → {FAIL_MARKER}")
        return FAIL_MARKER

    # ── Keyed batch wrappers (parallel mode — pinned key) ────────────────────

    def _translate_batch_keyed(self, batch, key, context_map, b_num, total_b):
        """Like _translate_batch but uses a pinned key instead of rotator."""
        job      = self.job
        mode     = job.get("op_mode", "translate")
        expected = [rn for rn, _ in batch]

        ctx_rows = None
        if mode == "translate_context" and context_map:
            ctx_rows = [(rn, context_map.get(rn, "")) for rn, _ in batch]

        prompt = build_translate_prompt(
            batch,
            src_lang=job["src_lang"],
            tgt_lang=job["tgt_lang"],
            style_prompt=job["style_prompt"],
            context_rows=ctx_rows,
            global_context=job.get("global_context"),
            custom_placeholders=job.get("custom_placeholders", ""),
        )

        raw = self._call_with_retry_keyed(prompt, key, f"Batch {b_num}/{total_b}",
                                          use_repair_model=False)
        if raw is None:
            with self._stats_lock:
                self.stats["failed"] += len(batch)
            return {rn: FAIL_MARKER for rn, _ in batch}

        parsed = parse_batch_response(raw, expected)
        result = {}

        for rn, src_text in batch:
            val = parsed.get(rn, "").strip()
            if val:
                result[rn] = val
                with self._print_lock:
                    print(f"  {GY}Row {rn:>5}{R}  {C2}{truncate(src_text,32)}{R}"
                          f"  {GY}→{R}  {GR}{truncate(val,32)}{R}")
                with self._stats_lock:
                    self.stats["ok"] += 1
            else:
                solo = self._solo_translate(rn, src_text, context_map, b_num, total_b)
                result[rn] = solo
                with self._stats_lock:
                    if solo == FAIL_MARKER:
                        self.stats["failed"] += 1
                    else:
                        self.stats["ok"] += 1
        return result

    def _repair_batch_keyed(self, rows_data, key, repair_mode, b_num, total_b):
        """Like _repair_batch but uses a pinned key instead of rotator."""
        expected = [d["row_num"] for d in rows_data]
        prompt   = build_repair_prompt(rows_data, repair_mode,
                                        src_lang=self.job.get("src_lang", ""),
                                        tgt_lang=self.job.get("tgt_lang", ""))

        raw = self._call_with_retry_keyed(prompt, key, f"Repair {b_num}/{total_b}",
                                          use_repair_model=True)
        if raw is None:
            with self._stats_lock:
                self.stats["failed"] += len(rows_data)
            return {d["row_num"]: FAIL_MARKER for d in rows_data}

        parsed = parse_batch_response(raw, expected)
        result = {}

        for d in rows_data:
            rn  = d["row_num"]
            val = parsed.get(rn, "").strip()
            if val:
                changed = "(repaired)" if val != d["translated"] else "(no change)"
                with self._print_lock:
                    print(f"  {GY}Row {rn:>5}{R}  {C2}{truncate(d['translated'],30)}{R}"
                          f"  {GY}→{R}  {MG}{truncate(val,30)}{R}  {GY}{changed}{R}")
                result[rn] = val
                with self._stats_lock:
                    self.stats["repaired"] += 1
                    self.stats["ok"]       += 1
            else:
                solo = self._solo_repair(d, repair_mode)
                result[rn] = solo
                with self._stats_lock:
                    if solo == FAIL_MARKER:
                        self.stats["failed"] += 1
                    else:
                        self.stats["repaired"] += 1
                        self.stats["ok"]       += 1
        return result

    def _call_with_retry_keyed(self, prompt, key, label, use_repair_model=False):
        """Pinned-key variant — same retry logic as _call_with_retry."""
        active_model  = self.repair_model if use_repair_model else self.model
        retry_mode    = self.job.get("retry_mode", "cooldown")
        max_attempts  = self.job.get("max_attempts", MAX_RETRIES)
        err_delay     = self.job.get("retry_delay_err", RETRY_DELAY_ERR)
        srv_delay     = self.job.get("retry_delay_503", RETRY_DELAY_503)
        temperature   = self.job.get("temperature", 0.1)
        infinite      = (retry_mode == "infinite")
        attempt       = 0

        while True:
            attempt += 1
            if not infinite and attempt > max_attempts:
                break

            t0 = time.time()
            try:
                response = call_llm(self.provider, key, active_model, prompt,
                                    temperature=temperature)
                elapsed  = time.time() - t0
                self.rotator.record_result(key, True, elapsed)
                return response

            except ProviderError as e:
                elapsed = time.time() - t0
                self.rotator.record_result(key, False, elapsed)
                is_503  = (e.status_code == 503)
                attempt_str = f"attempt {attempt}" if not infinite else f"attempt {attempt}/∞"

                if e.is_rate_limit:
                    with self._print_lock:
                        log_warn(f"[{label}] {attempt_str}: key ...{key[-4:]} rate limit — cooldown {e.retry_delay}s")
                    self._countdown(min(e.retry_delay, 90))
                elif is_503:
                    with self._print_lock:
                        log_warn(f"[{label}] {attempt_str}: 503 server error — waiting {srv_delay}s")
                    self._countdown(srv_delay)
                else:
                    with self._print_lock:
                        log_err(f"[{label}] {attempt_str}: {e}")
                    self._countdown(err_delay)

        with self._print_lock:
            log_fail(f"[{label}] key ...{key[-4:]} failed after {max_attempts} attempts.")
        return None

    # ── Core LLM call with retry + key rotation ───────────────────────────────

    def _call_with_retry(self, prompt, label, use_repair_model=False):
        """
        Retry logic respects job retry_mode:
          'cooldown'  — retry up to max_attempts, then give up
          'infinite'  — retry forever until success or Ctrl+C
        """
        active_model  = self.repair_model if use_repair_model else self.model
        retry_mode    = self.job.get("retry_mode", "cooldown")
        max_attempts  = self.job.get("max_attempts", MAX_RETRIES)
        err_delay     = self.job.get("retry_delay_err", RETRY_DELAY_ERR)
        srv_delay     = self.job.get("retry_delay_503", RETRY_DELAY_503)
        temperature   = self.job.get("temperature", 0.1)
        infinite      = (retry_mode == "infinite")
        attempt       = 0

        while True:
            attempt += 1
            if not infinite and attempt > max_attempts:
                break

            key, wait = self.rotator.get_key()
            if wait > 0:
                log_warn(f"All keys in cooldown — waiting {wait:.0f}s")
                self._countdown(int(wait) + 1)
                key, _ = self.rotator.get_key()

            t0 = time.time()
            try:
                response = call_llm(self.provider, key, active_model, prompt,
                                    temperature=temperature)
                elapsed  = time.time() - t0
                self.rotator.record_result(key, True, elapsed)
                return response

            except ProviderError as e:
                elapsed = time.time() - t0
                self.rotator.record_result(key, False, elapsed)
                is_503  = (e.status_code == 503)
                attempt_str = f"attempt {attempt}" if not infinite else f"attempt {attempt}/∞"

                if e.is_rate_limit:
                    self.rotator.block(key, e.retry_delay)
                    log_warn(f"[{label}] {attempt_str}: rate limit ...{key[-4:]} — cooldown {e.retry_delay}s")
                    self._countdown(min(e.retry_delay, 90))
                elif is_503:
                    log_warn(f"[{label}] {attempt_str}: 503 server error — waiting {srv_delay}s")
                    self._countdown(srv_delay)
                else:
                    log_err(f"[{label}] {attempt_str}: {e}")
                    self._countdown(err_delay)

        log_fail(f"[{label}] failed after {max_attempts} attempts.")
        return None

    def _countdown(self, seconds):
        for r in range(int(seconds), 0, -1):
            print(f"\r  {YL}Waiting {r:>3}s ...{R}  ", end="", flush=True)
            time.sleep(1)
        print(f"\r{' ' * 25}\r", end="", flush=True)


# ══════════════════════════════════════════════════════════════════════════════
#  WORKFLOW MANAGER
# ══════════════════════════════════════════════════════════════════════════════

class WorkflowManager:

    def __init__(self):
        self.session = {
            "provider_idx":  None,
            "keys":          [],
            "model":         None,   # translate model
            "repair_model":  None,   # repair model (None = same as model)
        }
        self.job = {}

    def _provider(self):
        return PROVIDERS[self.session["provider_idx"]]

    # ── Vault helpers ──────────────────────────────────────────────────────────

    def _vault_load_menu(self):
        if not keys_file_exists():
            log_warn("No vault file found.")
            return None

        for attempt in range(1, 4):
            pw = nav_input(f"Master password (attempt {attempt}/3): ")
            if pw in ("BACK", "CANCEL"):
                return pw
            keys = _load_keys(self._provider()["key"], pw)
            if keys is None:
                log_err("Wrong password.")
                if attempt == 3:
                    log_warn("Too many failed attempts.")
                    return None
            elif not keys:
                log_warn(f"No keys saved for {self._provider()['label']}.")
                return None
            else:
                for i, k in enumerate(keys, 1):
                    log_ok(f"Key #{i} loaded: {mask_key(k)}")
                return keys
        return None

    def _vault_manage(self):
        if not keys_file_exists():
            log_warn("No vault file found.")
            return

        pw = nav_input("Master password: ")
        if pw in ("BACK", "CANCEL"):
            return

        data = _load_all_raw(pw)
        if data is None:
            log_err("Wrong password.")
            return

        while True:
            print()
            print_menu(
                ["Show all saved keys", "Add keys to provider",
                 "Delete provider keys", "Change master password", "Back"],
                title="Vault Manager"
            )
            choice = nav_choose(
                ["show", "add", "delete", "change", "back"],
                "Select: ", allow_back=True
            )
            if choice in ("BACK", "CANCEL", 4):
                break

            if choice == 0:
                print()
                for prov in PROVIDERS:
                    keys = data.get(prov["key"], [])
                    masked = ", ".join(mask_key(k) for k in keys) if keys else "(none)"
                    print(f"  {GY}{prov['label']:<16}{R}  {WH}{masked}{R}")
                print()

            elif choice == 1:
                labels = [p["label"] for p in PROVIDERS]
                print_menu(labels, title="Select provider:")
                pi = nav_choose(PROVIDERS, "Provider: ")
                if pi in ("BACK", "CANCEL"):
                    continue
                pkey  = PROVIDERS[pi]["key"]
                new_keys = data.get(pkey, [])
                idx = len(new_keys) + 1
                while True:
                    k = nav_input(f"Key #{idx} (empty = done): ", allow_empty=True)
                    if k in ("BACK", "CANCEL") or k == "":
                        break
                    new_keys.append(k)
                    log_ok(f"Added: {mask_key(k)}")
                    idx += 1
                data[pkey] = new_keys
                _save_vault(data, pw)
                log_ok("Vault updated.")

            elif choice == 2:
                labels = [p["label"] for p in PROVIDERS]
                print_menu(labels, title="Select provider to delete:")
                pi = nav_choose(PROVIDERS, "Provider: ")
                if pi in ("BACK", "CANCEL"):
                    continue
                pkey = PROVIDERS[pi]["key"]
                if not _confirm(f"Delete all keys for {PROVIDERS[pi]['label']}? [y/N]: "):
                    continue
                data.pop(pkey, None)
                _save_vault(data, pw)
                log_ok("Keys deleted.")

            elif choice == 3:
                new_pw  = nav_input("New password: ")
                if new_pw in ("BACK", "CANCEL"):
                    continue
                confirm_pw = nav_input("Confirm new password: ")
                if confirm_pw in ("BACK", "CANCEL"):
                    continue
                if new_pw != confirm_pw:
                    log_err("Passwords do not match.")
                    continue
                _save_vault(data, new_pw)
                pw = new_pw
                log_ok("Password updated.")

    # ════════════════════════════════════════════════════════
    #  STEP FUNCTIONS
    # ════════════════════════════════════════════════════════

    def step_provider(self):
        print_step("Select Provider")
        labels = [p["label"] for p in PROVIDERS]
        print_menu(labels, title="Available providers:")
        print(f"  {GY}Type 'cancel' to exit{R}\n")

        result = nav_choose(PROVIDERS, "Provider [1-3]: ", allow_back=False)
        if result in ("BACK", "CANCEL"):
            return result
        self.session["provider_idx"] = result
        log_ok(f"Provider: {PROVIDERS[result]['label']}")
        return "OK"

    def step_apikeys(self):
        print_step("API Key Management")
        p = self._provider()
        log_info(f"Provider: {p['label']}   Key format: {p['api_key_hint']}")

        vault_status = "found" if keys_file_exists() else "not found"
        log_info(f"Vault: {CONFIG_FILE}  [{vault_status}]")
        print()

        print_menu(
            ["Load from vault", "Input manually", "Manage vault"],
            title="Options:"
        )
        choice = nav_choose(["load", "manual", "manage"], "Select [1-3]: ")
        if choice in ("BACK", "CANCEL"):
            return choice

        if choice == 0:
            keys = self._vault_load_menu()
            if keys in ("BACK", "CANCEL"):
                return keys
            if not keys:
                return "BACK"
            self.session["keys"] = keys
            log_ok(f"{len(keys)} key(s) loaded from vault.")
            return "OK"

        elif choice == 1:
            print()
            log_info("Enter keys one per line. Leave empty to finish.")
            log_info("Keys are visible while typing, masked after Enter.")
            print()
            keys = []
            idx  = 1
            while True:
                raw = nav_input(f"Key #{idx} (empty = done): ", allow_empty=True)
                if raw == "BACK":
                    if keys:
                        break
                    return "BACK"
                if raw == "CANCEL":
                    return "CANCEL"
                if raw == "":
                    if not keys:
                        log_err("Enter at least one API key.")
                        continue
                    break
                keys.append(raw)
                log_ok(f"Added: {mask_key(raw)}")
                idx += 1

            print()
            if _confirm("Save keys to vault? [y/N]: "):
                pw = nav_input("Master password for vault: ")
                if pw not in ("BACK", "CANCEL"):
                    ok = _save_keys(p["key"], keys, pw)
                    if ok:
                        log_ok("Keys saved to vault.")
                    else:
                        log_err("Wrong password — keys NOT saved.")

            self.session["keys"] = keys
            log_ok(f"{len(keys)} key(s) ready.")
            return "OK"

        else:
            self._vault_manage()
            return "BACK"

    def step_model(self):
        """Select the TRANSLATE model."""
        print_step("Select Translate Model")
        p   = self._provider()
        key = self.session["keys"][0]

        log_info(f"Fetching models from {p['label']} ...")
        try:
            models = fetch_models(p["key"], key)
        except ProviderError as e:
            log_err(f"Failed to fetch models: {e}")
            log_err("Check your API key and internet connection.")
            return "BACK"

        if not models:
            log_err("No models found.")
            return "BACK"

        log_ok(f"{len(models)} model(s) found.\n")
        print_menu(models, title="Available models:")
        print(f"  {GY}Type 'back' to go back{R}\n")

        result = nav_choose(models, f"Select translate model [1-{len(models)}]: ")
        if result in ("BACK", "CANCEL"):
            return result

        self.session["model"] = models[result]
        # Reset repair_model when translate model changes
        self.session["repair_model"] = None
        log_ok(f"Translate model: {self.session['model']}")
        return "OK"

    def step_repair_model(self):
        """
        Select the REPAIR model.
        Only shown when:
          - op_mode == 'translate' or 'translate_context' AND with_repair == True
          - op_mode == 'repair'
        """
        op_mode     = self.job.get("op_mode", "translate")
        with_repair = self.job.get("with_repair", False)

        # Only applicable when repair is involved
        if op_mode in ("translate", "translate_context") and not with_repair:
            return "SKIP"

        print_step("Select Repair Model")
        p   = self._provider()
        key = self.session["keys"][0]

        translate_model = self.session.get("model", "-")
        log_info(f"Translate model : {translate_model}")
        log_info(f"Choose a separate model for the Repair phase,")
        log_info(f"or keep the same model.")
        print()

        print_menu(
            [f"Use same model  ({translate_model})",
             "Choose a different model"],
            title="Repair model options:"
        )

        choice = nav_choose(["same", "different"], "Select [1-2]: ")
        if choice in ("BACK", "CANCEL"):
            return choice

        if choice == 0:
            # Same model
            self.session["repair_model"] = self.session["model"]
            log_ok(f"Repair model: {self.session['repair_model']}  (same as translate)")
            return "OK"

        # choice == 1: choose different model
        log_info(f"Fetching models from {p['label']} ...")
        try:
            models = fetch_models(p["key"], key)
        except ProviderError as e:
            log_err(f"Failed to fetch models: {e}")
            return "BACK"

        if not models:
            log_err("No models found.")
            return "BACK"

        log_ok(f"{len(models)} model(s) found.\n")
        print_menu(models, title="Available models:")
        print(f"  {GY}Current translate model: {WH}{translate_model}{R}\n")

        result = nav_choose(models, f"Select repair model [1-{len(models)}]: ")
        if result in ("BACK", "CANCEL"):
            return result

        self.session["repair_model"] = models[result]
        log_ok(f"Repair model: {self.session['repair_model']}")
        return "OK"

    def step_op_mode(self):
        print_step("Operation Mode")
        modes = [
            "Translate",
            "Translate with Context",
            "Repair Mode",
        ]
        descs = [
            "AI translates source text to target language.",
            "AI uses context column(s) to improve translation accuracy.",
            "AI repairs existing translated text (no new translation).",
        ]
        for i, (m, d) in enumerate(zip(modes, descs), 1):
            print(f"  {GY}[{i}]{R}  {WH}{m}{R}")
            print(f"       {GY}{d}{R}")
        print()

        result = nav_choose(modes, "Mode [1-3]: ")
        if result in ("BACK", "CANCEL"):
            return result

        self.job["op_mode"] = ["translate", "translate_context", "repair"][result]
        log_ok(f"Mode: {modes[result]}")
        return "OK"

    def step_mode_options(self):
        mode = self.job.get("op_mode", "translate")
        if mode in ("translate", "translate_context"):
            return self._translate_options(mode)
        else:
            return self._repair_type()

    def _translate_options(self, mode):
        print_step("Translate Options")

        if mode == "translate_context":
            log_info("Context helps AI understand abbreviations, keys, or domain-specific terms.")
            print()

            ctx_col_raw = nav_input(
                "Context column per-row [e.g. A2, leave empty to skip]: ",
                allow_empty=True
            )
            if ctx_col_raw in ("BACK", "CANCEL"):
                return ctx_col_raw
            if ctx_col_raw:
                col, row = parse_cell_ref(ctx_col_raw)
                if col is None:
                    log_err("Invalid cell reference.")
                    return "BACK"
                self.job["ctx_col"] = col
            else:
                self.job["ctx_col"] = None

            glob_ctx = nav_input(
                "Global context text (empty = none): ",
                allow_empty=True
            )
            if glob_ctx in ("BACK", "CANCEL"):
                return glob_ctx
            self.job["global_context"] = glob_ctx or None

            if not self.job.get("ctx_col") and not self.job.get("global_context"):
                log_err("At least one context source is required for this mode.")
                return "BACK"

        print()
        self.job["with_repair"] = _confirm("Enable Repair after Translate? [y/N]: ")
        if self.job["with_repair"]:
            log_ok("Translate + Repair enabled.")
            # Ask repair type immediately
            result = self._repair_type(title="Repair Type (after Translate)")
            if result in ("BACK", "CANCEL"):
                return result
        else:
            self.job["repair_mode"] = None
            log_ok("Translate only.")
        return "OK"

    def _repair_type(self, title="Repair Mode"):
        print_step(title)
        options = [
            ("repair",                "Repair Text",
             "Fix typos and grammar in the translated text only."),
            ("repair_compare",        "Repair + Compare Original",
             "Compare with original column to fix meaning inaccuracies."),
            ("repair_compare_context","Repair + Compare + Context",
             "Use original + context column for most accurate repair."),
        ]
        for i, (_, label, desc) in enumerate(options, 1):
            print(f"  {GY}[{i}]{R}  {WH}{label}{R}")
            print(f"       {GY}{desc}{R}")
        print()

        result = nav_choose(options, f"Repair type [1-{len(options)}]: ")
        if result in ("BACK", "CANCEL"):
            return result

        self.job["repair_mode"] = options[result][0]
        log_ok(f"Repair type: {options[result][1]}")
        return "OK"

    def step_file(self):
        print_step("File Configuration")
        xlsx_files = scan_excel_files(".")

        if xlsx_files:
            log_info("Excel files in current directory:")
            print_menu(xlsx_files)
            log_info("Enter number or full path.")
        else:
            log_warn("No .xlsx files found in current directory.")
            log_info("Enter the full path to the Excel file.")
        print()

        while True:
            raw = nav_input("Source file: ")
            if raw in ("BACK", "CANCEL"):
                return raw
            if xlsx_files:
                try:
                    idx = int(raw) - 1
                    if 0 <= idx < len(xlsx_files):
                        raw = xlsx_files[idx]
                except ValueError:
                    pass
            path = raw if os.path.isabs(raw) else os.path.join(".", raw)
            if not os.path.isfile(path):
                log_err(f"File not found: {path}")
                continue
            self.job["src_path"] = path
            log_ok(f"Source: {path}")
            break

        base        = os.path.splitext(os.path.basename(self.job["src_path"]))[0]
        op_mode     = self.job.get("op_mode", "translate")
        default_out = f"{base}_translated.xlsx" if op_mode != "repair" else f"{base}_repaired.xlsx"

        print()
        log_info(f"Default output filename: {default_out}")
        raw = nav_input(f"Output filename [{default_out}]: ", allow_empty=True)
        if raw in ("BACK", "CANCEL"):
            return raw
        out_name = raw if raw else default_out
        if not out_name.lower().endswith(".xlsx"):
            out_name += ".xlsx"

        out_path = os.path.join(os.path.dirname(self.job["src_path"]), out_name)
        try:
            shutil.copy2(self.job["src_path"], out_path)
        except Exception as e:
            log_err(f"Failed to create output file: {e}")
            return "BACK"

        self.job["out_path"] = out_path
        log_ok(f"Output: {out_path}")
        return "OK"

    def step_columns(self):
        print_step("Column Configuration")
        log_info("Format: column letter + start row  (e.g. C2, D2, AB5)")
        print()

        mode        = self.job.get("op_mode", "translate")
        repair_mode = self.job.get("repair_mode", "")
        with_repair = self.job.get("with_repair", False)

        prompts = []

        repair_needs_original = repair_mode in ("repair_compare", "repair_compare_context")
        repair_needs_context  = repair_mode == "repair_compare_context"

        if mode == "translate":
            prompts = [
                ("src_col",        "Source column        [e.g. C2]: "),
                ("translated_col",  "Translated column   [e.g. D2]: "),
            ]
            if with_repair:
                prompts.append(("repair_col", "Repair column        [e.g. E2]: "))

        elif mode == "translate_context":
            prompts = [
                ("src_col",        "Source column        [e.g. C2]: "),
                ("translated_col",  "Translated column   [e.g. D2]: "),
            ]
            if with_repair:
                prompts.append(("repair_col", "Repair column        [e.g. E2]: "))

        elif mode == "repair":
            if repair_mode == "repair":
                # Only needs translated col + repair output col
                prompts = [
                    ("translated_col", "Translated column    [e.g. D2]: "),
                    ("repair_col",      "Repair output column [e.g. E2]: "),
                ]
            elif repair_mode == "repair_compare":
                prompts = [
                    ("src_col",        "Original column      [e.g. C2]: "),
                    ("translated_col",  "Translated column   [e.g. D2]: "),
                    ("repair_col",      "Repair column       [e.g. E2]: "),
                ]
            else:  # repair_compare_context
                prompts = [
                    ("ctx_col",        "Context column       [e.g. A2]: "),
                    ("src_col",        "Original column      [e.g. C2]: "),
                    ("translated_col",  "Translated column   [e.g. D2]: "),
                    ("repair_col",      "Repair column       [e.g. E2]: "),
                ]

        start_row = None
        for key, prompt in prompts:
            while True:
                raw = nav_input(prompt)
                if raw in ("BACK", "CANCEL"):
                    return raw
                col, row = parse_cell_ref(raw)
                if col is None:
                    log_err("Invalid format. Use e.g. C2 or AB5.")
                    continue
                if start_row is None:
                    start_row = row
                self.job[key] = col
                log_ok(f"{prompt.strip().split('[')[0].strip()}: "
                       f"col {get_column_letter(col)}, start row {row}")
                break

        if start_row:
            self.job["start_row"] = start_row
        return "OK"

    def step_language(self):
        if self.job.get("op_mode") == "repair":
            return "SKIP"

        print_step("Language Configuration")
        log_info("Examples: Indonesian, English, Japanese, Korean, French")
        print()

        src = nav_input("Source language: ")
        if src in ("BACK", "CANCEL"):
            return src

        tgt = nav_input("Target language: ")
        if tgt in ("BACK", "CANCEL"):
            return tgt

        self.job["src_lang"] = src
        self.job["tgt_lang"] = tgt
        log_ok(f"{src}  →  {tgt}")
        return "OK"

    def step_preset(self):
        if self.job.get("op_mode") == "repair":
            return "SKIP"

        print_step("Translation Style")
        labels = [p["label"] for p in PRESETS]
        print_menu(labels, title="Available presets:")

        result = nav_choose(PRESETS, f"Select preset [1-{len(PRESETS)}]: ")
        if result in ("BACK", "CANCEL"):
            return result

        chosen = PRESETS[result]
        self.job["preset_key"] = chosen["key"]

        if chosen["key"] == "custom":
            prompt = nav_input("Enter custom prompt: ")
            if prompt in ("BACK", "CANCEL"):
                return prompt
            self.job["style_prompt"] = prompt
        else:
            self.job["style_prompt"] = chosen["prompt"]

        self.job["custom_placeholders"] = ""
        if "game" in chosen["key"]:
            cp = nav_input("Additional placeholders to preserve (empty = none): ",
                           allow_empty=True)
            if cp not in ("BACK", "CANCEL") and cp:
                self.job["custom_placeholders"] = cp

        log_ok(f"Preset: {chosen['label']}")
        return "OK"

    def step_batch_config(self):
        print_step("Row & Batch Configuration")

        while True:
            raw = nav_input("Number of rows to process: ")
            if raw in ("BACK", "CANCEL"):
                return raw
            try:
                n = int(raw)
                if n >= 1:
                    self.job["num_rows"] = n
                    break
            except ValueError:
                pass
            log_err("Enter a positive integer.")

        log_ok(f"Rows: {self.job['start_row']} → {self.job['start_row'] + n - 1}  ({n} rows)")
        print()

        while True:
            raw = nav_input("Rows per batch [default 10, max 100]: ")
            if raw in ("BACK", "CANCEL"):
                return raw
            if not raw:
                raw = "10"
            try:
                bs = int(raw)
                if 1 <= bs <= 100:
                    self.job["batch_size"] = bs
                    break
            except ValueError:
                pass
            log_err("Enter a number between 1 and 100.")

        if self.job["batch_size"] > 30:
            log_warn(f"Batch size {self.job['batch_size']} may reduce parse accuracy.")
            log_warn("Recommended: 10-30 for best results.")
            if not _confirm("Continue anyway? [y/N]: "):
                self.job["batch_size"] = 10
                log_ok("Batch size reset to 10.")

        log_ok(f"Batch size: {self.job['batch_size']} rows/request")
        print()

        # ── Inter-batch delay — Enter = 0.5s default ──────────────────────────
        while True:
            raw = nav_input(
                "Delay between batches in seconds [Enter = 0.5]: ",
                allow_empty=True          # ← allow empty so Enter works
            )
            if raw == "BACK":
                return "BACK"
            if raw == "CANCEL":
                return "CANCEL"
            if raw == "":
                raw = "0.5"               # ← default
            try:
                d = float(raw)
                if d >= 0:
                    self.job["inter_delay"] = d
                    break
            except ValueError:
                pass
            log_err("Enter a non-negative number.")

        log_ok(f"Inter-batch delay: {self.job['inter_delay']}s")
        print()

        self.job["skip_filled"] = _confirm(
            "Skip rows with existing output? (resume mode) [Y/n]: "
        )
        if self.job["skip_filled"]:
            log_ok("Resume mode ON — existing output rows will be skipped.")
        else:
            log_warn("Resume mode OFF — all rows will be overwritten.")
        return "OK"

    def step_threads(self):
        """Multi-thread per key config — runs before step_algorithm."""
        print_step("Multi-Thread per Key")
        n_keys = len(self.session["keys"])
        print(f"  {GY}Setiap API key bisa menjalankan beberapa thread sekaligus.{R}")
        print(f"  {GY}Semakin banyak thread → AI fokus ke lebih sedikit baris per request.{R}")
        print(f"  {GY}Max aman: 4 thread/key  |  {n_keys} key(s) loaded{R}")
        print()

        while True:
            raw = nav_input("Threads per key [Enter=1]: ", allow_empty=True)
            if raw == "BACK":  return "BACK"
            if raw == "CANCEL": return "CANCEL"
            raw = raw or "1"
            try:
                t = int(raw)
                if 1 <= t <= 10:
                    self.job["threads_per_key"] = t
                    break
            except ValueError:
                pass
            log_err("Enter a number between 1 and 10.")

        tpk      = self.job["threads_per_key"]
        total_w  = n_keys * tpk
        bs       = self.job.get("batch_size", 10)
        rows     = self.job.get("num_rows", 0)
        per_req  = max(1, bs // tpk) if tpk > 1 else bs
        log_ok(f"Threads per key : {tpk}")
        log_ok(f"Total workers   : {n_keys} key(s) × {tpk} thread(s) = {total_w} concurrent")
        if rows:
            effective_batch = max(1, rows // total_w) if total_w <= rows else 1
            log_ok(f"~{effective_batch} row(s) per AI request  ({rows} rows ÷ {total_w} workers)")
        print()
        return "OK"

    def step_retry_mode(self):
        print_step("Error Retry Mode")
        print(f"  {GY}[1]{R}  {WH}Cooldown{R}")
        print(f"       {GY}Retry N kali dengan delay, lalu stop jika tetap gagal.{R}")
        print(f"  {GY}[2]{R}  {WH}Infinite Retry{R}")
        print(f"       {GY}Retry terus sampai berhasil. Hanya berhenti jika Ctrl+C.{R}")
        print()

        raw = nav_input("Retry mode [1-2, Enter=1]: ", allow_empty=True)
        if raw == "BACK":   return "BACK"
        if raw == "CANCEL": return "CANCEL"
        raw = raw or "1"

        if raw == "2":
            self.job["retry_mode"] = "infinite"
            print()
            while True:
                r = nav_input("Delay antar attempt (detik) [Enter=10]: ", allow_empty=True)
                if r in ("BACK","CANCEL"): return r
                r = r or "10"
                try:
                    d = float(r)
                    if d >= 0:
                        self.job["retry_delay_err"] = d
                        self.job["retry_delay_503"] = max(d, 30)
                        break
                except ValueError: pass
                log_err("Enter a non-negative number.")
            log_ok(f"Infinite retry  — delay {self.job['retry_delay_err']}s  |  503 delay {self.job['retry_delay_503']}s")
        else:
            self.job["retry_mode"] = "cooldown"
            print()
            while True:
                r = nav_input("Max attempts [Enter=3]: ", allow_empty=True)
                if r in ("BACK","CANCEL"): return r
                r = r or "3"
                try:
                    v = int(r)
                    if v >= 1:
                        self.job["max_attempts"] = v
                        break
                except ValueError: pass
                log_err("Enter a positive integer.")

            while True:
                r = nav_input("Delay error biasa (detik) [Enter=5]: ", allow_empty=True)
                if r in ("BACK","CANCEL"): return r
                r = r or "5"
                try:
                    d = float(r)
                    if d >= 0:
                        self.job["retry_delay_err"] = d
                        break
                except ValueError: pass
                log_err("Enter a non-negative number.")

            while True:
                r = nav_input("Delay khusus 503 (detik) [Enter=45]: ", allow_empty=True)
                if r in ("BACK","CANCEL"): return r
                r = r or "45"
                try:
                    d = float(r)
                    if d >= 0:
                        self.job["retry_delay_503"] = d
                        break
                except ValueError: pass
                log_err("Enter a non-negative number.")

            log_ok(f"Cooldown — {self.job['max_attempts']}x attempts  |  "
                   f"err delay {self.job['retry_delay_err']}s  |  503 delay {self.job['retry_delay_503']}s")
        print()
        return "OK"

    def step_temperature(self):
        print_step("Model Temperature")
        print(f"  {GY}Rendah (0.1) = konsisten & akurat   |   Tinggi (0.7+) = kreatif/bervariasi{R}")
        print(f"  {GY}Recommended untuk translate: 0.1{R}")
        print()
        while True:
            raw = nav_input("Temperature [Enter=0.1]: ", allow_empty=True)
            if raw == "BACK":   return "BACK"
            if raw == "CANCEL": return "CANCEL"
            raw = raw or "0.1"
            try:
                t = float(raw)
                if 0.0 <= t <= 2.0:
                    self.job["temperature"] = t
                    break
            except ValueError:
                pass
            log_err("Enter a number between 0.0 and 2.0.")
        log_ok(f"Temperature: {self.job['temperature']}")
        print()
        return "OK"

    def step_debug_log(self):
        print_step("Debug Log")
        print(f"  {GY}Tampilkan raw AI response saat parse error.{R}")
        print(f"  {GY}Berguna untuk diagnosa masalah format output AI.{R}")
        print()
        enabled = _confirm("Enable debug log? [y/N]: ")
        self.job["debug_log"] = enabled
        if enabled:
            log_ok("Debug log: ON — raw response akan ditampilkan saat parse error.")
        else:
            log_ok("Debug log: OFF")
        print()
        return "OK"

    def step_algorithm(self):
        print_step("API Key Distribution Algorithm")
        n_keys = len(self.session["keys"])
        log_info(f"{n_keys} key(s) loaded.")
        print()

        algos = [
            ("roundrobin", "Round Robin",
             "Keys used in rotation. Simple and predictable."),
            ("weighted",   "Weighted Queue",
             "Assign weight per key. Higher weight = more batches."),
            ("adaptive",   "Adaptive Dynamic Queue",
             "Auto-selects best key based on success rate & speed."),
            ("parallel",   "Parallel",
             f"Batches divided across ALL {n_keys} key(s) and run simultaneously. "
             f"Fastest option — {n_keys} batches per wave."),
        ]
        for i, (_, label, desc) in enumerate(algos, 1):
            print(f"  {GY}[{i}]{R}  {WH}{label}{R}")
            print(f"       {GY}{desc}{R}")
        print()

        result = nav_choose(algos, f"Algorithm [1-{len(algos)}]: ")
        if result in ("BACK", "CANCEL"):
            return result

        self.job["algorithm"]   = algos[result][0]
        self.job["key_weights"] = None
        self.job["n_workers"]   = None
        log_ok(f"Algorithm: {algos[result][1]}")

        if self.job["algorithm"] == "weighted":
            print()
            log_info("Assign weight 1-10 per key (higher = more batches).")
            weights = []
            for i, k in enumerate(self.session["keys"], 1):
                while True:
                    raw = nav_input(f"Weight for key #{i} ({mask_key(k)}) [default 5]: ",
                                    allow_empty=True)
                    if raw in ("BACK", "CANCEL"):
                        return raw
                    raw = raw or "5"
                    try:
                        w = int(raw)
                        if 1 <= w <= 10:
                            weights.append(w)
                            break
                    except ValueError:
                        pass
                    log_err("Enter a number between 1 and 10.")
            self.job["key_weights"] = weights
            log_ok(f"Weights: {weights}")

        elif self.job["algorithm"] == "parallel":
            print()
            log_info(f"Workers = jumlah API key yang digunakan per wave.")
            log_info(f"Default = {n_keys} (semua key).")
            print()
            while True:
                raw = nav_input(
                    f"Jumlah workers [Enter = {n_keys} (semua key)]: ",
                    allow_empty=True
                )
                if raw == "BACK":
                    return "BACK"
                if raw == "CANCEL":
                    return "CANCEL"
                if raw == "":
                    self.job["n_workers"] = n_keys
                    break
                try:
                    w = int(raw)
                    if 1 <= w <= n_keys:
                        self.job["n_workers"] = w
                        break
                    log_err(f"Enter a number between 1 and {n_keys}.")
                except ValueError:
                    log_err("Enter a valid number.")
            log_ok(f"Workers: {self.job['n_workers']} parallel  "
                   f"({self.job['n_workers']} batches per wave)")

        return "OK"

    def step_preview(self):
        print_step("Preview & Confirm")
        job  = self.job
        mode = job.get("op_mode", "translate")

        col_labels = []
        if mode == "repair":
            rm = job.get("repair_mode", "repair")
            if "context" in rm and job.get("ctx_col"):
                col_labels.append((job["ctx_col"], "Context"))
            if "compare" in rm and job.get("src_col"):
                col_labels.append((job["src_col"], "Original"))
            if job.get("translated_col"):
                label = "Translated"
                col_labels.append((job["translated_col"], label))
            if job.get("repair_col"):
                col_labels.append((job["repair_col"], "Repair (output)"))
        else:
            if job.get("ctx_col"):
                col_labels.append((job["ctx_col"], "Context"))
            if job.get("src_col"):
                col_labels.append((job["src_col"], "Source"))
            if job.get("translated_col"):
                col_labels.append((job["translated_col"], "Translated (output)"))
            if job.get("repair_col") and job.get("with_repair"):
                col_labels.append((job["repair_col"], "Repair (output)"))

        try:
            n_prev  = min(3, job.get("num_rows", 3))
            preview = preview_rows(job["src_path"], col_labels,
                                   job["start_row"], n_prev)
        except Exception as e:
            log_err(f"Failed to read file: {e}")
            return "BACK"

        _div()
        print(f"  {B}{WH}Data Preview  (first {n_prev} rows):{R}\n")
        hdr = f"  {'Row':<5}" + "".join(f"  {lbl:<22}" for _, lbl in col_labels)
        print(f"{GY}{hdr}{R}")
        _div()
        for row in preview:
            line = f"  {WH}{row['_row']:<5}{R}"
            for _, lbl in col_labels:
                line += f"  {C2}{truncate(row.get(lbl,''), 22):<22}{R}"
            print(line)
        _div()
        print()

        # Summary
        n_workers   = job.get("n_workers", 1)
        algo_labels = {"roundrobin": "Round Robin", "weighted": "Weighted Queue",
                       "adaptive":   "Adaptive Dynamic Queue",
                       "parallel":   f"Parallel ({n_workers} workers/wave)"}
        mode_labels = {"translate": "Translate",
                       "translate_context": "Translate with Context",
                       "repair": "Repair Mode"}
        repair_labels = {

            "repair":                 "Repair Text",
            "repair_compare":         "Repair + Compare Original",
            "repair_compare_context": "Repair + Compare + Context",
        }

        total_b     = max(1, job.get("num_rows", 0) // max(1, job.get("batch_size", 10)))
        with_repair = job.get("with_repair", False)
        api_calls   = total_b * 2 if with_repair else total_b

        translate_model = self.session.get("model", "-")
        repair_model    = self.session.get("repair_model") or translate_model
        same_model      = translate_model == repair_model

        rows = [
            ("Provider",    self._provider()["label"]),
            ("Model",       translate_model),
        ]

        # Show repair model row only when repair is involved
        if with_repair or job.get("op_mode") == "repair":
            repair_label = repair_model + (" (same)" if same_model else " ← different")
            rows.append(("Repair model", repair_label))

        rows.append(("Keys",        f"{len(self.session['keys'])} key(s)"))
        rows.append(("Mode",        mode_labels.get(mode, mode)))

        if mode == "repair":
            rows.append(("Repair type", repair_labels.get(job.get("repair_mode",""), "")))
        else:
            rows.append(("Language",    f"{job.get('src_lang','?')} → {job.get('tgt_lang','?')}"))
            if job.get("preset_key"):
                label = next((p["label"] for p in PRESETS
                              if p["key"] == job["preset_key"]), "")
                rows.append(("Style", label))
            rows.append(("Repair",  "Enabled" if with_repair else "Disabled"))

        rows += [
            ("Source",      job.get("src_path", "-")),
            ("Output",      job.get("out_path", "-")),
            ("Rows",        f"{job.get('start_row','?')} → "
                            f"{job.get('start_row',0) + job.get('num_rows',0) - 1}"
                            f"  ({job.get('num_rows',0)} rows)"),
            ("Batch size",  f"{job.get('batch_size',10)} rows/batch  →  ~{total_b} batches"),
            ("API calls",   f"~{api_calls}"),
            ("Algorithm",   algo_labels.get(job.get("algorithm",""), "")),
            ("Skip filled", "Yes (resume)" if job.get("skip_filled") else "No (overwrite)"),
        ]

        print(f"  {B}Job Summary:{R}\n")
        for label, val in rows:
            print(f"  {GY}{label:<14}{R}  {WH}{val}{R}")
        print()
        _div()
        print()

        if not _confirm("Start processing? [Y/n]: "):
            return "BACK"
        return "OK"

    def step_process(self):
        print_step("Processing")

        job      = self.job
        mode     = job.get("op_mode", "translate")
        rotator  = build_rotator(
            self.session["keys"],
            job.get("algorithm", "adaptive"),
            job.get("key_weights"),
        )

        translate_model  = self.session["model"]
        repair_model     = self.session.get("repair_model") or translate_model
        algorithm        = job.get("algorithm", "adaptive")
        n_keys           = len(rotator.all_keys())
        threads_per_key  = job.get("threads_per_key", 1)
        # n_workers: for parallel = keys × threads; for others = just threads
        if algorithm == "parallel":
            n_workers = job.get("n_workers", n_keys) * threads_per_key
        else:
            n_workers = threads_per_key
        is_parallel = (algorithm == "parallel") or (threads_per_key > 1)

        algo_label = {
            "roundrobin": "Round Robin",
            "weighted":   "Weighted Queue",
            "adaptive":   "Adaptive Dynamic Queue",
            "parallel":   "Parallel",
        }.get(algorithm, algorithm)
        log_info(f"Algorithm      : {algo_label}  |  Keys: {n_keys}  |  Threads/key: {threads_per_key}")
        log_info(f"Total workers  : {n_workers} concurrent requests")
        log_info(f"Translate model: {translate_model}  |  temp={job.get('temperature',0.1)}")
        if job.get("with_repair") or mode == "repair":
            same = "(same)" if translate_model == repair_model else ""
            log_info(f"Repair model   : {repair_model}  {same}")
        log_info(f"Retry mode     : {job.get('retry_mode','cooldown')}  |  "
                 f"503 delay={job.get('retry_delay_503', RETRY_DELAY_503)}s")
        if is_parallel and job.get("num_rows"):
            total_b = max(1, (job["num_rows"] + job.get("batch_size",10) - 1) // job.get("batch_size",10))
            waves   = max(1, (total_b + n_workers - 1) // n_workers)
            log_info(f"Wave size      : {n_workers} batches/wave  →  ~{waves} wave(s)")
        _div()
        print()

        engine = ProcessingEngine(
            provider_key=self._provider()["key"],
            rotator=rotator,
            model=translate_model,
            job=job,
            repair_model=repair_model,
        )

        t_start = time.time()

        try:
            if mode in ("translate", "translate_context"):
                source_rows = load_column(
                    job["src_path"], job["src_col"],
                    job["start_row"], job["num_rows"]
                )
                context_map = None
                if mode == "translate_context" and job.get("ctx_col"):
                    ctx_rows    = load_column(
                        job["src_path"], job["ctx_col"],
                        job["start_row"], job["num_rows"]
                    )
                    context_map = {rn: txt for rn, txt in ctx_rows}

                if is_parallel:
                    engine.run_translate_parallel(source_rows, context_map, n_workers)
                else:
                    engine.run_translate(source_rows, context_map)

            else:   # repair mode
                rm          = job.get("repair_mode", "repair")
                trans_rows  = load_column(
                    job["out_path"], job["translated_col"],
                    job["start_row"], job["num_rows"]
                )

                orig_map = {}
                ctx_map  = {}
                if "compare" in rm and job.get("src_col"):
                    orig_rows = load_column(
                        job["out_path"], job["src_col"],
                        job["start_row"], job["num_rows"]
                    )
                    orig_map = {r: t for r, t in orig_rows}
                if "context" in rm and job.get("ctx_col"):
                    ctx_rows = load_column(
                        job["out_path"], job["ctx_col"],
                        job["start_row"], job["num_rows"]
                    )
                    ctx_map = {r: t for r, t in ctx_rows}

                rows_data = []
                for rn, txt in trans_rows:
                    d = {"row_num": rn, "translated": txt}
                    if orig_map:
                        d["original"] = orig_map.get(rn, "")
                    if ctx_map:
                        d["context"] = ctx_map.get(rn, "")
                    rows_data.append(d)

                if is_parallel:
                    engine.run_repair_parallel(rows_data, n_workers)
                else:
                    engine.run_repair(rows_data)

        except Exception as e:
            log_err(f"Processing error: {e}")
            return "BACK"

        job["stats"]   = engine.stats
        job["elapsed"] = time.time() - t_start
        return "OK"

    def step_summary(self):
        print_step("Summary")
        stats   = self.job.get("stats", {})
        elapsed = self.job.get("elapsed", 0)
        m, s    = divmod(int(elapsed), 60)
        t_str   = f"{m}m {s}s" if m else f"{s}s"

        _div("═")
        print()
        print(f"  {B}{GR}Processing complete.{R}\n")

        rows = [
            (GR,  "OK",       stats.get("ok",      0)),
            (MG,  "Repaired", stats.get("repaired", 0)),
            (GY,  "Skipped",  stats.get("skipped",  0)),
            (RD,  "Failed",   stats.get("failed",   0)),
            (WH,  "Total",    stats.get("total",    0)),
        ]
        for color, label, val in rows:
            bar = f"{GY}{'█' * min(val, 28)}{R}"
            print(f"  {color}{label:<12}{R}  {color}{val:>5}{R}  {bar}")

        print(f"\n  {GY}Time    :{R}  {WH}{t_str}{R}")
        print(f"  {GY}Output  :{R}  {WH}{self.job.get('out_path', '-')}{R}")
        print()
        _div("═")
        return "OK"

    # ════════════════════════════════════════════════════════
    #  MAIN RUN LOOP
    # ════════════════════════════════════════════════════════

    def run(self):
        print_banner()
        first_run = True

        while True:
            self.job = {}

            if first_run or not self.session.get("keys"):
                start_at = "provider"
            else:
                p = self._provider()
                log_info(
                    f"Active session: {p['label']} / "
                    f"{self.session['model']} / "
                    f"{len(self.session['keys'])} key(s)"
                )
                start_at = "op_mode" if _confirm(
                    "Reuse provider / model / keys? [Y/n]: "
                ) else "provider"

            first_run = False

            all_steps = [
                ("provider",      self.step_provider),
                ("apikeys",       self.step_apikeys),
                ("model",         self.step_model),
                ("op_mode",       self.step_op_mode),
                ("mode_opts",     self.step_mode_options),
                ("repair_model",  self.step_repair_model),
                ("file",          self.step_file),
                ("columns",       self.step_columns),
                ("language",      self.step_language),
                ("preset",        self.step_preset),
                ("batch",         self.step_batch_config),
                ("threads",       self.step_threads),
                ("retry_mode",    self.step_retry_mode),
                ("temperature",   self.step_temperature),
                ("debug_log",     self.step_debug_log),
                ("algorithm",     self.step_algorithm),
                ("preview",       self.step_preview),
                ("process",       self.step_process),
                ("summary",       self.step_summary),
            ]

            start_idx = next(
                (i for i, (sid, _) in enumerate(all_steps) if sid == start_at), 0
            )

            stack = []
            i     = start_idx

            while i < len(all_steps):
                step_id, step_fn = all_steps[i]
                result = step_fn()

                if result == "SKIP":
                    i += 1

                elif result == "OK":
                    stack.append(i)
                    i += 1

                elif result == "BACK":
                    if stack:
                        i = stack.pop()
                    elif i > start_idx:
                        i = start_idx
                    else:
                        log_warn("Already at first step.")
                        i += 1

                elif result == "CANCEL":
                    sys.exit(0)

                else:
                    i += 1

            print()
            if not _confirm("Process another file? [y/N]: "):
                print(f"\n  {GR}Thank you for using Universal Excel Translator.{R}")
                print(f"  {GY}Powered by Gemini · Groq · Cerebras{R}\n")
                break


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    def _sigint(sig, frame):
        print(f"\n\n  {YL}Interrupted. Goodbye.{R}\n")
        sys.exit(0)
    signal.signal(signal.SIGINT, _sigint)

    WorkflowManager().run()


if __name__ == "__main__":
    main()
